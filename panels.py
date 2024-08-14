import json
import os
import pathlib
import shutil
import tkinter as tk
import datetime as dt
import customtkinter as ctk
from tkinter import ttk
from tkinter import (messagebox,
                     filedialog)
from typing import (List,
                    Dict, )
from base_classes import (MyFrame,
                          MyScrolledText,
                          MyLabel,
                          MyEntry,
                          MyButton,
                          MyListbox,
                          MyInputDialog,
                          MyRadiobutton,
                          MyDialog)
from lang_module import LANG
from zoomed_canvas import ZoomAdvanced
from my_logging_module import (my_logger,
                               log_exception,
                               log_exception_decorator)
from openpyxl import Workbook


@log_exception_decorator(log_exception)
class TreePan(MyFrame):
    """TreePan is placed in top left corner.
    """
    def __init__(self, *args, **kwargs) -> None:
        super(TreePan, self).__init__(*args, **kwargs)

        self.rowconfigure(0, weight=1)
        self.rowconfigure(1, weight=0)
        self.columnconfigure(0, weight=1)
        self.columnconfigure(1, weight=0)

        self.tree = ttk.Treeview(master=self, show="headings")

        self.style = ttk.Style()
        self.style.configure("Treeview.Heading", font=(None, 18))
        self.style.configure("Treeview", font=(None, 18), rowheight=28)

        columns = ("f_name", "source", "path", "labels", "c_time")
        self.tree.configure(columns=columns)

        self.tree.heading("f_name", text=LANG.get("f_name"))
        self.tree.heading("source", text=LANG.get("source"))
        self.tree.heading("path", text=LANG.get("path"))
        self.tree.heading("labels", text=LANG.get("labels"))
        self.tree.heading("c_time", text=LANG.get("c_time"))

        self.tree.grid(row=0, column=0, sticky=tk.NSEW)

        scroll_y = ctk.CTkScrollbar(master=self, orientation=tk.VERTICAL, command=self.tree.yview)
        scroll_y.grid(row=0, column=1, sticky=tk.NS)
        self.tree.configure(yscrollcommand=scroll_y.set)

        scroll_x = ctk.CTkScrollbar(master=self, orientation=tk.HORIZONTAL, command=self.tree.xview)
        scroll_x.grid(row=1, column=0, sticky=tk.EW)
        self.tree.configure(xscrollcommand=scroll_x.set)

        self.tree.bind('<<TreeviewSelect>>', self.selecting_item)

        self.populate()
        my_logger.debug("TreePan initiated")

    def populate(self, filtered_file_ids=None) -> None:
        """Important note: tree iids = uuids.
        Method support filtering as well.
        """
        self.tree.delete(*self.tree.get_children())

        if not filtered_file_ids:
            for _uuid, data in self.master.brain.project["files"].items():
                self.tree.insert(
                    "",
                    tk.END,
                    iid=_uuid,
                    values=(
                        data.get("f_name"),
                        data.get("source"),
                        data.get("path"),
                        data.get("labels"), # todo (prawodpodobnie tutaj i niżej wpadają {})
                        f"{dt.datetime.fromtimestamp(int(data.get('c_time'))):%Y-%m-%d %H:%M:%S}"
                    )
                )
        else:
            for _uuid in filtered_file_ids:
                data = self.master.brain.project["files"][_uuid]
                self.tree.insert(
                    "",
                    tk.END,
                    iid=_uuid,
                    values=(
                        data.get("f_name"),
                        data.get("source"),
                        data.get("path"),
                        data.get("labels"),
                        f"{dt.datetime.fromtimestamp(int(data.get('c_time'))):%Y-%m-%d %H:%M:%S}"
                    )
                )

        my_logger.debug("TreePan.populate - done.")

    def selecting_item(self, event) -> None:
        """Important note: tree iids = uuids"""
        item_index_uuid = self.tree.focus()

        my_logger.debug(f"TreePan.selecting item for {item_index_uuid}")

        self.master.brain.mount_detail_and_zoom_pan(_uuid=item_index_uuid)


@log_exception_decorator(log_exception)
class ZoomPan(MyFrame):
    """Object placed in top right corner"""
    def __init__(self, path_to_img=None, *args, **kwargs) -> None:
        super(ZoomPan, self).__init__(*args, **kwargs)

        if not path_to_img:
            return

        self.zoom = ZoomAdvanced(master=self, path=path_to_img)
        self.zoom.grid(sticky=tk.NSEW)


@log_exception_decorator(log_exception)
class DetailPan(MyFrame):
    """Object placed in left bottom corner"""
    def __init__(self, *args, **kwargs) -> None:

        super(DetailPan, self).__init__(*args, **kwargs)

        self.columnconfigure(0, weight=0)
        self.columnconfigure(1, weight=1)
        self.columnconfigure(2, weight=0)

        for row in range(8):
            self.rowconfigure(row, weight=0)

        self._uuid: str = ""
        self.file_data: Dict | None = None

        self.f_name: str = LANG.get("choose_file")
        self.f_name_var = tk.StringVar(master=self)
        self.f_name_var.set(self.f_name)

        self.id: int = 0

        self.source: str = ""
        self.source_var = tk.StringVar(master=self)
        self.source_var.set("")

        self.path: str = ""
        self.labels: List = []
        self.comment: str = ""

        self.c_time: str = ""

        self.main_options_pan = MyFrame(master=self)
        self.main_options_pan.grid(row=0, column=0, columnspan=3)
        self.main_options_pan.rowconfigure(0, weight=0)
        self.main_options_pan.columnconfigure((0, 1, 2, 3), weight=1)

        self.btn_del_file = MyButton(master=self.main_options_pan, text=LANG.get("del_file"),
                                     command=self.del_file, state=tk.DISABLED, hover=True)
        self.btn_del_file.grid(row=0, column=0)  # todo command

        self.btn_open_in_default_viewer = MyButton(master=self.main_options_pan, text=LANG.get("open_default_viewer"),
                                                   command=self.open_in_default_viewer, state=tk.DISABLED, hover=True)
        self.btn_open_in_default_viewer.grid(row=0, column=1)

        self.btn_move_up_tree = MyButton(master=self.main_options_pan, text=LANG.get("move_up"),
                                         command=self.move_up_tree, state=tk.DISABLED, hover=True)
        self.btn_move_up_tree.grid(row=0, column=2)

        self.btn_move_down_tree = MyButton(master=self.main_options_pan, text=LANG.get("move_down"),
                                           command=self.move_down_tree, state=tk.DISABLED, hover=True)
        self.btn_move_down_tree.grid(row=0, column=3)

        MyLabel(master=self, text=LANG.get("f_name")).grid(row=1, column=0)

        self.ent_name = MyEntry(master=self, textvariable=self.f_name_var)
        self.ent_name.grid(row=1, column=1, sticky=tk.EW)

        self.btn_rename_file = MyButton(master=self, text=LANG.get("f_rename"),
                                        command=self.rename_file, state=tk.DISABLED, hover=True)
        self.btn_rename_file.grid(row=1, column=2)

        MyLabel(master=self, text=LANG.get("source")).grid(row=2, column=0)

        self.source_pan = MyFrame(master=self)
        self.source_pan.grid(row=2, column=1, columnspan=2)
        self.source_pan.rowconfigure((0, 1), weight=0)
        self.source_pan.columnconfigure((0, 1), weight=1)

        self.ent_src = MyEntry(master=self.source_pan, textvariable=self.source_var)
        self.ent_src.grid(row=0, column=0, columnspan=2, sticky=tk.EW)

        self.btn_rename_source = MyButton(master=self.source_pan, text=LANG.get("s_rename"),
                                          command=self.rename_source, state=tk.DISABLED, hover=True)
        self.btn_rename_source.grid(row=1, column=0)

        self.btn_open_source_in_browser = MyButton(master=self.source_pan, text=LANG.get("open_src_browser"),
                                                   command=self.open_source_in_browser, state=tk.DISABLED, hover=True)
        self.btn_open_source_in_browser.grid(row=1, column=1)

        MyLabel(master=self, text=LANG.get("path")).grid(row=3, column=0)

        self.path_var = tk.StringVar(master=self)
        self.path_var.set(self.path)

        self.lbl_path = MyLabel(master=self, textvariable=self.path_var, width=100)
        self.lbl_path.grid(row=3, column=1)

        MyLabel(master=self, text=LANG.get("labels")).grid(row=4, column=0)

        self.lbls_listbox = MyListbox(master=self)
        self.lbls_listbox.grid(row=4, column=1)

        self.frame_lbls_btns = MyFrame(master=self)
        self.frame_lbls_btns.grid(row=4, column=2)
        self.frame_lbls_btns.rowconfigure((0, 4), weight=1)
        self.frame_lbls_btns.rowconfigure((1, 2, 3,), weight=0)
        self.frame_lbls_btns.columnconfigure(0, weight=0)
        MyLabel(master=self.frame_lbls_btns, text=LANG.get("manage_labels")).grid(row=0, column=0)

        self.btn_add_labels_to_file = MyButton(master=self.frame_lbls_btns, text=LANG.get("add_labels"),
                                               command=self.add_labels_to_file, state=tk.DISABLED, hover=True) # todo
        self.btn_add_labels_to_file.grid(row=1, column=0, sticky=tk.S)

        self.btn_remove_labels = MyButton(master=self.frame_lbls_btns, text=LANG.get("rem_labels"),
                                          command=self.remove_labels, state=tk.DISABLED, hover=True) # todo
        self.btn_remove_labels.grid(row=2, column=0)

        self.btn_move_lbl_up = MyButton(master=self.frame_lbls_btns, text=LANG.get("move_up"),
                                        command=self.move_lbl_up, state=tk.DISABLED, hover=True) # todo
        self.btn_move_lbl_up.grid(row=3, column=0)

        self.btn_move_lbl_down = MyButton(master=self.frame_lbls_btns, text=LANG.get("move_down"),
                                          command=self.move_lbl_down, state=tk.DISABLED, hover=True) # todo
        self.btn_move_lbl_down.grid(row=4, column=0, sticky=tk.N)

        MyLabel(master=self, text=LANG.get("comment")).grid(row=5, column=0)

        self.tbox_com = MyScrolledText(master=self, height=4)
        self.tbox_com.grid(row=5, column=1, sticky=tk.EW)

        self.btn_alter_comment = MyButton(master=self, text=LANG.get("alter_comment"),
                                          command=self.alter_comment, state=tk.DISABLED, hover=True) # todo
        self.btn_alter_comment.grid(row=5, column=2)

        MyLabel(master=self, text=LANG.get("c_time")).grid(row=7, column=0)

        self.c_time_var = tk.StringVar(master=self)
        self.c_time_var.set("")

        lbl_ctime = MyLabel(master=self, textvariable=self.c_time_var, width=100)
        lbl_ctime.grid(row=7, column=1)

    def update_pan(self, _uuid: str) -> None:
        self._uuid = _uuid

        self.file_data = self.master.brain.project["files"][self._uuid]

        self.f_name = self.file_data["f_name"]
        self.f_name_var.set(self.f_name)

        self.id = self.file_data["index"]

        self.source = self.file_data["source"]
        self.source_var.set(self.source)

        self.path = self.file_data["path"]
        path_broken = self.path
        if len(self.path) > 100:
            path_broken = path_broken[:99] + "\n    " + path_broken[99:]
        self.path_var.set(path_broken)

        self.labels = self.file_data["labels"]
        self.lbls_listbox.delete(0, tk.END)
        self.lbls_listbox.insert(tk.END, *self.labels)

        self.comment = self.file_data["comment"]
        self.tbox_com.delete("1.0", tk.END)
        self.tbox_com.insert("1.0", self.comment)

        self.c_time = self.file_data["c_time"]
        c_time = f"{dt.datetime.fromtimestamp(int(self.c_time)):%Y-%m-%d %H:%M:%S}"

        self.c_time_var.set(c_time)

        self.master.brain.set_file_or_project_history()

        self.btn_del_file.configure(state=tk.NORMAL, hover=True)
        self.btn_open_in_default_viewer.configure(state=tk.NORMAL, hover=True)
        self.btn_move_up_tree.configure(state=tk.NORMAL, hover=True)
        self.btn_move_down_tree.configure(state=tk.NORMAL, hover=True)
        self.btn_rename_file.configure(state=tk.NORMAL, hover=True)
        self.btn_rename_source.configure(state=tk.NORMAL, hover=True)
        self.btn_open_source_in_browser.configure(state=tk.NORMAL, hover=True)
        self.btn_add_labels_to_file.configure(state=tk.NORMAL, hover=True)
        self.btn_remove_labels.configure(state=tk.NORMAL, hover=True)
        self.btn_move_lbl_up.configure(state=tk.NORMAL, hover=True)
        self.btn_move_lbl_down.configure(state=tk.NORMAL, hover=True)
        self.btn_alter_comment.configure(state=tk.NORMAL, hover=True)

        my_logger.debug(f"DetailPan.update_pan for {self._uuid}: {self.f_name}")

    def del_file(self) -> None:  # todo command
        self.master.brain.del_file(self._uuid)

    def open_in_default_viewer(self) -> None:
        self.master.brain.open_in_default_viewer(self.path)

    def move_up_tree(self) -> None:
        self.master.brain.move_up_tree()

    def move_down_tree(self) -> None:
        self.master.brain.move_down_tree()

    def rename_file(self) -> None:
        n_name = self.f_name_var.get()

        my_logger.debug(f"DetailPan.rename_file for {self._uuid}: {self.f_name}, new: {n_name} - initiated")

        if not n_name.endswith(pathlib.Path(self.path).suffix):
            messagebox.showerror(master=self, title=LANG.get("f_rename"), message=LANG.get("f_name_wrong_suffix"))
            return

        if n_name != self.f_name:
            res = messagebox.askyesno(master=self, title=LANG.get("f_rename"), message=LANG.get("f_rename"))
            if res:
                old_path = pathlib.Path(self.path)
                old_path_abs = self.master.brain.project_path.joinpath(old_path)
                new_path = pathlib.Path(old_path.parent, n_name)
                new_path_abs = pathlib.Path(old_path_abs.parent, n_name)

                old_path_abs.rename(new_path_abs)
                self.path = str(new_path)

                path_broken = self.path
                if len(self.path) > 100:
                    path_broken = path_broken[:99] + "\n    " + path_broken[99:]
                self.lbl_path.configure(text=path_broken)

                self.f_name = n_name
                self.f_name_var.set(self.f_name)

                self.master.brain.rename_file(
                    self._uuid,
                    self.f_name,
                    self.path
                )

        else:
            messagebox.showerror(master=self, title=LANG.get("f_rename"), message=LANG.get("same_f_name"))

    def rename_source(self) -> None:
        src = self.source_var.get()

        my_logger.debug(f"DetailPan.rename_source for {self._uuid}:, old: {self.source}, new: {src}")

        if src != self.source:
            res = messagebox.askyesno(master=self, title=LANG.get("s_rename"), message=LANG.get("s_rename"))
            if res:
                self.source = src
                self.source_var.set(src)

                self.master.brain.rename_source(
                    self._uuid,
                    self.source
                )

                # self.master.brain.history_manager.save_previous_state(self.uuid, "path", self.path)
        else:
            messagebox.showerror(master=self, title=LANG.get("s_rename"), message=LANG.get("same_source"))

    def open_source_in_browser(self) -> None:  # moved to th brain
        self.master.brain.open_source_in_browser(self.source_var.get())

    # def update_fields(self, key: str, value: str | List) -> None:
    #     if key == "labels":
    #         self.lbls_listbox.delete(0, tk.END)
    #         self.lbls_listbox.insert(tk.END, *value.split(" /// "))
    #     elif key == "comment":
    #         self.tbox_com.delete("1.0", tk.END)
    #         self.tbox_com.insert("1.0", value)
    #     else:
    #         ...
    #
    #     my_logger.debug(f"DetailPan.update_fields: {self.file_id}: {key}: {value} - updated")

    def add_labels_to_file(self) -> None:
        self.master.brain.add_labels_to_file(self)

    def remove_labels(self) -> None:
        indexes = self.lbls_listbox.curselection()
        if indexes:
            labels_to_remove = [self.lbls_listbox.get(index) for index in indexes]
            for index in indexes:
                self.lbls_listbox.delete(index)

            self.labels = [label for label in self.labels if label not in labels_to_remove]

            my_logger.debug(f"DetailPan {self._uuid}: Labels to remove: {labels_to_remove}")
            my_logger.debug(f"DetailPan {self._uuid}: Labels left: {self.labels}")

            self.master.brain.move_or_remove_file_labels(
                self._uuid,
                self.labels
            )

    def move_lbl_up(self) -> None:
        indexes = self.lbls_listbox.curselection()
        if indexes:
            index = indexes[0]
            if index == 0:
                return

            curr_label = self.lbls_listbox.get(index)
            del self.labels[index]
            self.labels.insert(index - 1, curr_label)
            self.lbls_listbox.delete(index)
            self.lbls_listbox.insert(index - 1, curr_label)

            self.lbls_listbox.selection_set(index - 1)

            my_logger.debug(f"DetailPan {self._uuid}: Label moved up: {curr_label}")
            my_logger.debug(f"DetailPan {self._uuid}: Current labels: {self.labels}")

            self.master.brain.move_or_remove_file_labels(
                self._uuid,
                self.labels
            )

    def move_lbl_down(self) -> None:
        indexes = self.lbls_listbox.curselection()
        if indexes:
            index = indexes[0]
            if index == len(self.labels) - 1:
                return

            curr_label = self.lbls_listbox.get(index)
            self.labels.insert(index + 2, curr_label)
            del self.labels[index]
            self.lbls_listbox.insert(index + 2, curr_label)
            self.lbls_listbox.delete(index)

            self.lbls_listbox.selection_set(index + 1)

            my_logger.debug(f"DetailPan {self._uuid}: Label moved down: {curr_label}")
            my_logger.debug(f"DetailPan {self._uuid}: Current labels: {self.labels}")

            self.master.brain.move_or_remove_file_labels(
                self._uuid,
                self.labels
            )

    def alter_comment(self) -> None:
        comment = self.tbox_com.get("1.0", tk.END)[:-1]

        my_logger.debug(f"DetailPan.alter_comment for {self._uuid}, old: {self.comment}, new: {comment}")

        if comment != self.comment:
            res = messagebox.askyesno(master=self, title=LANG.get("alter_comment"), message=LANG.get("alter_comment"))
            if res:
                self.master.brain.alter_comment(self._uuid, comment)
                self.comment = comment
        else:
            messagebox.showerror(master=self, title=LANG.get("alter_comment"), message=LANG.get("same_comment"))


@log_exception_decorator(log_exception)
class RotatingPan(MyFrame):
    """Object placed in the bottom right corner.
    This is a frame for LabelPan, ProjectHistoryPan, FileHistoryPan, FilterPan, LookUpPan
    """
    def __init__(self, *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)

        self.rowconfigure(0, weight=0)
        self.rowconfigure(1, weight=1)

        self.columnconfigure(0, weight=1)

        self.menu_bar = MyFrame(master=self)
        self.menu_bar.grid(row=0, column=0)

        self.menu_bar.rowconfigure(0, weight=0)
        self.menu_bar.columnconfigure((0, 1, 2, 3, 4), weight=1)

        self.pan = MyFrame(master=self)
        self.pan.grid(row=1, column=0)

        MyButton(master=self.menu_bar, text=LANG.get("file_history"), command=self.file_history_pan).grid(row=0, column=0)
        MyButton(master=self.menu_bar, text=LANG.get("proj_hist"), command=self.project_history_pan).grid(row=0, column=1)
        MyButton(master=self.menu_bar, text=LANG.get("lbl_pan"), command=self.label_pan).grid(row=0, column=2)
        MyButton(master=self.menu_bar, text=LANG.get("filter_pan"), command=self.filter_pan).grid(row=0, column=3)
        MyButton(master=self.menu_bar, text=LANG.get("look_up"), command=self.look_up_pan).grid(row=0, column=4)

        my_logger.debug("RotatingPan initialized")

    def file_history_pan(self) -> None:
        self.pan.destroy()
        self.pan = FileHistoryPan(master=self)
        self.pan.grid(row=1, column=0)

        self.master.brain.set_file_or_project_history()

    def project_history_pan(self) -> None:
        self.pan.destroy()
        self.pan = ProjectHistoryPan(master=self)
        self.pan.grid(row=1, column=0)

        self.master.brain.set_file_or_project_history()

    def label_pan(self) -> None:
        self.pan.destroy()
        self.pan = LabelPan(master=self)
        self.pan.grid(row=1, column=0)

    def filter_pan(self) -> None:
        self.pan.destroy()
        self.pan = FilterPan(master=self)
        self.pan.grid(row=1, column=0)

    def look_up_pan(self) -> None:
        self.pan.destroy()
        self.pan = LookUpPan(master=self)
        self.pan.grid(row=1, column=0)


@log_exception_decorator(log_exception)
class LabelPan(MyFrame):
    def __init__(self, *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)

        self.rowconfigure(0, weight=0)
        self.rowconfigure(1, weight=1)
        # self.rowconfigure(2, weight=1)

        self.columnconfigure(0, weight=1)
        self.columnconfigure(1, weight=0)

        MyLabel(master=self, text=LANG.get("lbls_in_pro")).grid(row=0, column=0)

        self.all_lbls_listbox = MyListbox(master=self, selectmode=tk.MULTIPLE)
        self.all_lbls_listbox.grid(row=1, column=0)

        self.all_lbls_listbox.insert(tk.END, *self.master.master.brain.project["labels"])

        self.button_pan = MyFrame(master=self)
        self.button_pan.grid(row=1, column=1)
        self.button_pan.columnconfigure(0, weight=0)
        self.button_pan.rowconfigure((0, 4), weight=1)
        self.button_pan.rowconfigure((1, 2, 3,), weight=0)

        MyButton(master=self.button_pan, text=LANG.get("move_up"), command=self.move_up).grid(row=0, column=0, sticky=tk.S)
        MyButton(master=self.button_pan, text=LANG.get("move_down"), command=self.move_down).grid(row=1, column=0)
        MyButton(master=self.button_pan, text=LANG.get("add_lbl_to_project"), command=self.add_label_to_project).grid(
            row=2, column=0)
        MyButton(master=self.button_pan, text=LANG.get("remane_lbl"), command=self.rename_label).grid(row=3, column=0)
        MyButton(master=self.button_pan, text=LANG.get("del_lbl"), command=self.delete_label).grid(row=4, column=0, sticky=tk.N)

        my_logger.debug("LabelPan initialized")

    def move_up(self) -> None:
        my_logger.debug("LabelPan.move_up")
        index = self.all_lbls_listbox.curselection()
        if not index:
            return
        index = index[0]
        selected = self.all_lbls_listbox.get(index)
        if index - 1 < 0:
            return

        all_labels = list(self.all_lbls_listbox.get(0, tk.END))
        all_labels.remove(selected)
        all_labels.insert(index - 1, selected)

        self.all_lbls_listbox.delete(0, tk.END)
        self.all_lbls_listbox.insert(tk.END, *all_labels)
        self.all_lbls_listbox.selection_set(index - 1)

        self.master.master.brain.save_new_order_all_labels(all_labels)

    def move_down(self) -> None:
        my_logger.debug("LabelPan.move_down")
        index = self.all_lbls_listbox.curselection()
        if not index:
            return
        index = index[0]
        selected = self.all_lbls_listbox.get(index)
        all_labels = list(self.all_lbls_listbox.get(0, tk.END))
        if index + 1 > len(all_labels):
            return

        all_labels.remove(selected)
        all_labels.insert(index + 1, selected)

        self.all_lbls_listbox.delete(0, tk.END)
        self.all_lbls_listbox.insert(tk.END, *all_labels)

        self.all_lbls_listbox.selection_set(index + 1)

        self.master.master.brain.save_new_order_all_labels(all_labels)

    def rename_label(self) -> None:
        my_logger.debug("LabelPan.rename_label")
        index = self.all_lbls_listbox.curselection()
        if not index:
            return
        index = index[0]
        old_label = self.all_lbls_listbox.get(index)

        dialog = MyInputDialog(text=LANG.get("type_lbl_name"), title=LANG.get("add_lbl_to_pro"))
        new_label = dialog.get_input()

        if new_label:
            all_labels = list(self.all_lbls_listbox.get(0, tk.END))

            all_labels.insert(index, new_label)
            all_labels.remove(old_label)
            self.all_lbls_listbox.delete(0, tk.END)
            self.all_lbls_listbox.insert(tk.END, *all_labels)

            self.master.master.brain.rename_label(old_label, new_label, all_labels)

    def delete_label(self) -> None:
        my_logger.debug("LabelPan.delete_label")
        index = self.all_lbls_listbox.curselection()
        if not index:
            return
        index = index[0]
        label = self.all_lbls_listbox.get(index)

        msg_title_message = LANG.get("msg_message_delete") + label

        res = messagebox.askyesno(master=self, title=msg_title_message,
                                  message=msg_title_message)
        if res:
            self.all_lbls_listbox.delete(index)
            self.master.master.brain.delete_label(label)

    def add_label_to_project(self) -> None:
        my_logger.debug("LabelPan.add_label_to_project")
        dialog = MyInputDialog(text=LANG.get("no_new_label_name"), title=LANG.get("add_label"))
        new_label = dialog.get_input()
        if not new_label:
            return
        if new_label in self.all_lbls_listbox.get(0, tk.END):
            return

        self.all_lbls_listbox.insert(tk.END, new_label)
        self.master.master.brain.add_label_to_project(new_label)


@log_exception_decorator(log_exception)
class ProjectHistoryPan(MyFrame):
    def __init__(self, *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)

        self.rowconfigure(0, weight=0)
        self.rowconfigure(1, weight=1)
        self.columnconfigure(0, weight=1)
        self.columnconfigure(1, weight=0)

        MyLabel(master=self, text=LANG.get("proj_hist")).grid(row=0, column=0)

        self.history_tree = ttk.Treeview(master=self, show="headings")
        self.history_tree.grid(row=1, column=0, sticky=tk.NSEW)
        columns = ("record", )
        self.history_tree.configure(columns=columns)

        self.history_tree.heading("record", text=LANG.get("record"))

        MyButton(master=self, text=LANG.get("restore"), command=self.restore_for_project).grid(row=1, column=1)

        self.dict_translator = {}

        my_logger.debug(f"ProjectHistoryPan established")

    def set_project_history(self, project_history: Dict) -> None:
        my_logger.debug(f"ProjectHistoryPan.set_project_history started.")
        self.history_tree.delete(*self.history_tree.get_children())

        self.dict_translator = {}

        for key, data in project_history.items():
            self.dict_translator[key] = data

            # key, data = record
            f_name_or_project, _uuid_or_project_time_record = key.split(", ", 1)
            data = ", ".join([f"{key}: {val}" for key, val in data.items()])
            self.history_tree.insert(
                "",
                tk.END,
                iid=key,
                values=(f"{f_name_or_project}: {data}", )
            )

        my_logger.debug(f"ProjectHistoryPan.set_project_history: project history established")

    def restore_for_project(self) -> None:
        my_logger.debug("ProjectHistoryPan.restore_for_project")
        key = self.history_tree.focus()
        if key:
            data = self.dict_translator.get(key)

            self.master.master.brain.restore_previous_state(key, data)

            self.history_tree.delete(key)

            my_logger.debug(f"Restoring for project accomplished: {key}")


@log_exception_decorator(log_exception)
class FileHistoryPan(MyFrame):
    def __init__(self, *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)

        self.rowconfigure(0, weight=0)
        self.rowconfigure(1, weight=0)
        self.rowconfigure(2, weight=1)
        self.columnconfigure(0, weight=1)
        self.columnconfigure(1, weight=0)

        MyLabel(master=self, text=LANG.get("file_history")).grid(row=0, column=0)

        self._uuid = ""

        self.file_id_var = tk.StringVar()
        self.file_id_var.set("")

        MyLabel(master=self, textvariable=self.file_id_var).grid(row=1, column=0)

        self.history_tree = ttk.Treeview(master=self, show="headings")
        self.history_tree.grid(row=2, column=0, sticky=tk.NSEW)

        columns = ("record", )
        self.history_tree.configure(columns=columns)

        self.history_tree.heading("record", text=LANG.get("record"))

        MyButton(master=self, text=LANG.get("restore"), command=self.restore_for_file).grid(row=2, column=1)

        my_logger.debug(f"FileHistoryPan established")

        self.dict_translator = {}

    def set_file_history(self, _uuid: str, f_name: str, file_history: Dict) -> None:
        my_logger.debug("FileHistoryPan.set_file_history")
        self._uuid = _uuid

        self.file_id_var.set(f_name)

        self.history_tree.delete(*self.history_tree.get_children())

        self.dict_translator = {}

        for key, data in file_history.items():
            self.dict_translator[key] = data

            to_listbox = []
            for key_, data_ in data.items():
                to_listbox.append(f"{key_}: {data_}")
            self.history_tree.insert(
                "", tk.END,
                iid=key,
                values=(", ".join(to_listbox), )
            )

        my_logger.debug(f"FileHistoryPan.set_project_history: file history {f_name} established")

    def restore_for_file(self) -> None:
        my_logger.debug("FileHistoryPan.restore_for_file")
        key = self.history_tree.focus()
        if key:
            data = self.dict_translator.get(key)

            self.master.master.brain.restore_previous_state(key, data)

            # self.history_tree.delete(key)

            my_logger.debug(f"Restoring for {key} accomplished")


@log_exception_decorator(log_exception)
class FilterPan(MyFrame):
    def __init__(self, *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)
        self.rowconfigure(0, weight=0)
        self.rowconfigure(1, weight=0)
        self.rowconfigure(2, weight=0)
        self.rowconfigure(3, weight=0)
        self.rowconfigure(4, weight=1)
        self.rowconfigure(5, weight=0)
        self.rowconfigure(6, weight=0)
        self.rowconfigure(7, weight=0)
        self.rowconfigure(8, weight=1)
        self.rowconfigure(9, weight=0)
        self.columnconfigure(0, weight=1)

        MyLabel(master=self, text=LANG.get("filter_pan")).grid(row=0, column=0)
        MyLabel(master=self, text=LANG.get("sel_lbls_incl")).grid(row=1, column=0)
        # MyLabel(master=self, text="Logical operator:").grid(row=2, column=0)

        self.operator_labels_include = tk.BooleanVar()
        self.operator_labels_exclude = tk.BooleanVar()

        self.option_frame_include = MyFrame(master=self)
        self.option_frame_include.grid(row=3, column=0)

        self.option_frame_include.columnconfigure((0,1), weight=1)
        self.option_frame_include.rowconfigure(0, weight=0)

        MyRadiobutton(
            master=self.option_frame_include,
            variable=self.operator_labels_include,
            text=LANG.get("or_include"),
            value=False
        ).grid(column=0, row=0, sticky=tk.E)

        MyRadiobutton(
            master=self.option_frame_include,
            variable=self.operator_labels_include,
            text=LANG.get("and_include"),
            value=True
        ).grid(column=1, row=0, sticky=tk.W)

        self.lst_box_include = MyListbox(master=self, selectmode=tk.MULTIPLE, exportselection=False)
        self.lst_box_include.grid(row=4, column=0)

        labels = self.master.master.brain.project["labels"]

        self.lst_box_include.insert(tk.END, *labels)

        MyLabel(master=self, text=LANG.get("sel_lbls_excl")).grid(row=5, column=0)
        # MyLabel(master=self, text="Logical operator:").grid(row=6, column=0)

        self.option_frame_exclude = MyFrame(master=self)
        self.option_frame_exclude.grid(row=7, column=0)

        self.option_frame_exclude.columnconfigure((0,1), weight=1)
        self.option_frame_exclude.rowconfigure(0, weight=0)

        MyRadiobutton(
            master=self.option_frame_exclude,
            variable=self.operator_labels_exclude,
            text=LANG.get("or_exclude"),
            value=False
        ).grid(column=0, row=0, sticky=tk.E)

        MyRadiobutton(
            master=self.option_frame_exclude,
            variable=self.operator_labels_exclude,
            text=LANG.get("and_exclude"),
            value=True
        ).grid(column=1, row=0, sticky=tk.W)

        self.lst_box_exclude = MyListbox(master=self, selectmode=tk.MULTIPLE, exportselection=False)
        self.lst_box_exclude.grid(row=8, column=0)

        self.lst_box_exclude.insert(tk.END, *labels)

        self.choice_frame = MyFrame(master=self)
        self.choice_frame.grid(row=9, column=0)

        self.choice_frame.rowconfigure(0, weight=0)
        self.choice_frame.columnconfigure((0, 1), weight=1)

        MyButton(master=self.choice_frame, text=LANG.get("filter"), command=self.filter).grid(row=0, column=0)
        MyButton(master=self.choice_frame, text=LANG.get("show_all"), command=self.show_all).grid(row=0, column=1)

        my_logger.debug("FilterPan - established")

    def show_all(self) -> None:
        my_logger.debug("FilterPan.show_all")
        self.master.master.brain.reset_tree()

    def filter(self) -> None:
        positive = self.lst_box_include.curselection()
        negative = self.lst_box_exclude.curselection()

        my_logger.debug(f"FilterPan.filter, positive: {positive}, negative: {negative}")

        positive = [self.lst_box_include.get(p) for p in positive]
        negative = [self.lst_box_exclude.get(n) for n in negative]

        include = self.operator_labels_include.get()
        exclude = self.operator_labels_exclude.get()

        project = self.master.master.brain.project

        file_id_filtered_positive = []
        file_id_filtered_negative = []

        if include:
            for file_id, i in project["files"].items():
                lbls = i["labels"]
                p = [l for l in lbls if l in positive]
                if len(p) == len(positive):
                    file_id_filtered_positive.append(file_id)

        if exclude:
            for file_id, i in project["files"].items():
                lbls = i["labels"]
                n = [l for l in lbls if l in negative]
                if len(n) == len(negative):
                    file_id_filtered_negative.append(file_id)

        if not include:
            for file_id, i in project["files"].items():
                lbls = i["labels"]
                p = [l for l in lbls if l in positive]
                if p:
                    file_id_filtered_positive.append(file_id)

        if not exclude:
            for file_id, i in project["files"].items():
                lbls = i["labels"]
                n = [l for l in lbls if l in negative]
                if n:
                    file_id_filtered_negative.append(file_id)

        file_id_filtered = [f_id for f_id in file_id_filtered_positive if f_id not in file_id_filtered_negative]

        self.master.master.brain.filter_tree(file_id_filtered)


@log_exception_decorator(log_exception)
class LookUpPan(MyFrame):
    def __init__(self, *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)
        self.rowconfigure(0, weight=0)
        self.rowconfigure(1, weight=0)
        self.rowconfigure(2, weight=0)
        self.rowconfigure(3, weight=0)
        self.rowconfigure(4, weight=0)
        self.rowconfigure(5, weight=0)
        # self.rowconfigure(6, weight=0)
        self.columnconfigure(0, weight=0)
        self.columnconfigure(1, weight=1)
        self.columnconfigure(2, weight=1)

        MyLabel(master=self, text="Wyszukaj w aktualnie widocznych rekordach").grid(row=0, column=0, columnspan=3)

        MyLabel(master=self, text="Zawiera:").grid(row=1, column=1)
        MyLabel(master=self, text="Niezawiera:").grid(row=1, column=2)

        MyLabel(master=self, text="W nazwie").grid(row=2, column=0)

        self.name_include_var = tk.StringVar()
        self.entry_name_include = MyEntry(master=self, textvariable=self.name_include_var)
        self.entry_name_include.grid(row=2, column=1, sticky=tk.EW)

        self.name_exclude_var = tk.StringVar()
        self.ent_name_exclude = MyEntry(master=self, textvariable=self.name_exclude_var)
        self.ent_name_exclude.grid(row=2, column=2, sticky=tk.EW)

        MyLabel(master=self, text="W źródle").grid(row=3, column=0)

        self.source_include_var = tk.StringVar()
        self.ent_source_include = MyEntry(master=self, textvariable=self.source_include_var)
        self.ent_source_include.grid(row=3, column=1, sticky=tk.EW)

        self.source_exclude_var = tk.StringVar()
        self.ent_source_exclude = MyEntry(master=self, textvariable=self.source_exclude_var)
        self.ent_source_exclude.grid(row=3, column=2, sticky=tk.EW)

        MyLabel(master=self, text="W komentarzu").grid(row=4, column=0)

        self.comment_include_var = tk.StringVar()
        self.ent_comment_include = MyEntry(master=self, textvariable=self.comment_include_var)
        self.ent_comment_include.grid(row=4, column=1, sticky=tk.EW)

        self.comment_exclude_var = tk.StringVar()
        self.ent_comment_exclude = MyEntry(master=self, textvariable=self.comment_exclude_var)
        self.ent_comment_exclude.grid(row=4, column=2, sticky=tk.EW)

        MyButton(master=self, text="Wyszukaj", command=self.look_up).grid(row=5, column=0, columnspan=2)
        MyButton(master=self, text=LANG.get("show_all"), command=self.show_all).grid(row=5, column=3)

    def look_up(self):
        my_logger.debug("LookUpPan.look_up")
        to_include = [self.name_include_var.get(), self.source_include_var.get(), self.comment_include_var.get()]
        to_include = [x if x else None for x in to_include]

        to_exclude = [self.name_exclude_var.get(), self.source_exclude_var.get(), self.comment_exclude_var.get()]
        to_exclude = [x if x else None for x in to_exclude]

        my_logger.debug(f"LookUpPan.look_up: to_include: {to_include}, to_exclude: {to_exclude}")

        self.master.master.brain.look_up(to_include, to_exclude)

    def show_all(self) -> None:
        my_logger.debug("LookUpPan.show_all")
        self.master.master.brain.reset_tree()


@log_exception_decorator(log_exception)
class DistributeMan(ctk.CTkToplevel):
    """Object embedded in an independent window. I was too lazy to implement it inside MainWindow and export functions to TheBrain,
    but it works so mission accomplished :)
    """
    def __init__(self, *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)

        self.title(LANG.get("dist_manager"))

        self.grab_set()

        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=0)
        self.rowconfigure(1, weight=1)

        self.info_frame = MyFrame(master=self)
        self.info_frame.grid(column=0, row=0)

        self.info_frame.columnconfigure(0)
        self.info_frame.rowconfigure(0)

        info_text = LANG.get("info_text")

        MyLabel(master=self.info_frame, text=info_text, anchor=tk.W).grid(column=0, row=0)

        self.main_frame = MyFrame(master=self)
        self.main_frame.grid(column=0, row=1)

        self.main_frame.columnconfigure((0, 1, 2), weight=1)
        self.main_frame.rowconfigure(0, weight=1)

        self.selection_frame = MyFrame(master=self.main_frame)
        self.selection_frame.grid(column=0, row=0)

        self.selection_frame.columnconfigure(0, weight=1)
        self.selection_frame.rowconfigure(0, weight=0)
        self.selection_frame.rowconfigure(1, weight=1)
        self.selection_frame.rowconfigure(2)
        self.selection_frame.rowconfigure(3)
        self.selection_frame.rowconfigure(4)

        MyLabel(master=self.selection_frame, text=LANG.get("sel_labels")).grid(column=0, row=0)

        self.selection_listbox = MyListbox(master=self.selection_frame, selectmode=tk.MULTIPLE, exportselection=False)
        self.selection_listbox.grid(column=0, row=1)
        lbls = self.master.master.brain.project["labels"]
        self.selection_listbox.insert(tk.END, *lbls)

        MyButton(master=self.selection_frame, text=LANG.get("push_sel_labels"),
                 command=self.push_selected_label).grid(column=0, row=2)
        MyButton(master=self.selection_frame, text=LANG.get("save_schema"),
                 command=self.save_schema).grid(column=0, row=3)
        MyButton(master=self.selection_frame, text=LANG.get("load_schema"),
                 command=self.load_schema).grid(column=0, row=4)

        self.combined_frame = MyFrame(master=self.main_frame)
        self.combined_frame.grid(column=1, row=0)
        self.combined_frame.columnconfigure(0, weight=1)
        self.combined_frame.rowconfigure(0, weight=0)
        self.combined_frame.rowconfigure(1, weight=1)
        self.combined_frame.rowconfigure(2)
        self.combined_frame.rowconfigure(3)
        self.combined_frame.rowconfigure(4)

        MyLabel(master=self.combined_frame, text=LANG.get("comb_labels")).grid(column=0, row=0)

        self.combined_listbox = MyListbox(master=self.combined_frame, selectmode=tk.MULTIPLE, exportselection=False)
        self.combined_listbox.grid(column=0, row=1)

        MyButton(master=self.combined_frame, text=LANG.get("push_all_lbls_to_folders"),
                 command=self.push_all_combined_labels).grid(column=0, row=2)
        MyButton(master=self.combined_frame, text=LANG.get("push_lbl_to_folders"),
                 command=self.push_combined_label).grid(column=0, row=3)
        MyButton(master=self.combined_frame, text=LANG.get("del_comb"),
                 command=self.del_combined).grid(column=0, row=4)

        self.folder_frame = MyFrame(master=self.main_frame)
        self.folder_frame.grid(column=2, row=0)

        self.folder_frame.columnconfigure(0, weight=1)
        self.folder_frame.rowconfigure(0, weight=0)
        self.folder_frame.rowconfigure(1, weight=1)
        self.folder_frame.rowconfigure(2)
        self.folder_frame.rowconfigure(3)
        self.folder_frame.rowconfigure(4)

        MyLabel(master=self.folder_frame, text=LANG.get("folder")).grid(column=0, row=0)

        self.folder_listbox = MyListbox(master=self.folder_frame, selectmode=tk.MULTIPLE, exportselection=False)
        self.folder_listbox.grid(column=0, row=1)

        MyButton(master=self.folder_frame, text=LANG.get("rename_folder"),
                 command=self.rename_folder).grid(column=0, row=2)
        MyButton(master=self.folder_frame, text=LANG.get("prep_dist"),
                 command=self.prep_distribution).grid(column=0, row=3)
        MyButton(master=self.folder_frame, text=LANG.get("del_folder"),
                 command=self.del_folder).grid(column=0, row=4)

        self.folders_and_combinations = dict()

        self.distributor = None
        self.save_load = None

        my_logger.debug("DistributeMan initiated")

    def push_selected_label(self):
        selected = self.selection_listbox.curselection()
        if selected:
            selected = [self.selection_listbox.get(i) for i in selected]
            selected = " ___ ".join(selected)

            my_logger.debug(f"DistributeMan.push_selected_label:  {selected}")

            self.combined_listbox.insert(tk.END, selected)

    def save_schema(self):
        self.save_load = SaveLoadSchema(master=self, save=True)

    def load_schema(self):
        self.save_load = SaveLoadSchema(master=self, save=False)

    def push_all_combined_labels(self):
        all_comb_labels = self.combined_listbox.get(0, tk.END)

        my_logger.debug(f"DistributeMan.push_all_combined_labels: {all_comb_labels}")

        for comb_label in all_comb_labels:
            key_folder = comb_label

            if key_folder in self.folders_and_combinations:
                return

            self.folders_and_combinations[key_folder] = key_folder
            self.folder_listbox.insert(tk.END, key_folder)

    def push_combined_label(self):
        selected = self.combined_listbox.curselection()
        if selected:
            key_folder = self.combined_listbox.get(selected[0])

            if key_folder in self.folders_and_combinations:
                return

            my_logger.debug(f"DistributeMan.push_combined_label: {key_folder}")

            self.folders_and_combinations[key_folder] = key_folder
            self.folder_listbox.insert(tk.END, key_folder)

    def del_combined(self):
        selected = self.combined_listbox.curselection()
        if selected:
            item_in_folder = self.combined_listbox.get(selected[0])

            my_logger.debug(f"DistributeMan.del_combined: {item_in_folder}")

            for key, val in self.folders_and_combinations.items():
                if item_in_folder == val:
                    msg_message = LANG.get("operation_will_affect") + key
                    res = messagebox.askyesno(master=self, title=LANG.get("dist_manager"),
                                              message=msg_message)
                    if res:
                        self.combined_listbox.delete(selected[0])
                        self.folder_listbox.delete(0, tk.END)

                        my_logger.debug(f"DistributeMan.del_combined: proceeded to del: {key}")

                        del self.folders_and_combinations[key]

                        for key_, val_ in self.folders_and_combinations.items():
                            self.folder_listbox.insert(tk.END, val_)
                        break
                    break
            self.combined_listbox.delete(selected[0])

    def rename_folder(self):

        selected = self.folder_listbox.curselection()
        if selected:
            item_id = selected[0]
            old_name = self.folder_listbox.get(item_id)

            msg_title_message = LANG.get("rename_") + old_name

            new_folder_name = MyDialog(title=msg_title_message, text=msg_title_message)
            new_folder_name = new_folder_name.get_input()

            if new_folder_name:
                my_logger.debug(f"DistributeMan.rename_folder: old: {old_name}, new: {new_folder_name}")

                self.folders_and_combinations[new_folder_name] = self.folders_and_combinations[old_name]
                del self.folders_and_combinations[old_name]

                self.folder_listbox.delete(item_id)
                self.folder_listbox.insert(item_id, new_folder_name)

    def prep_distribution(self):
        my_logger.debug(f"DistributeMan.prep_distribution: initiated")
        if self.folders_and_combinations:
            self.distributor = Distributor(self.folders_and_combinations)
        else:
            messagebox.showinfo(master=self, title=LANG.get("dist_manager"), message=LANG.get("prep_folders_first"))

    def del_folder(self):
        selected = self.folder_listbox.curselection()
        if selected:
            folder = self.folder_listbox.get(selected[0])

            my_logger.debug(f"DistributeMan.del_folder: {folder}")

            del self.folders_and_combinations[folder]
            self.folder_listbox.delete(selected[0])


@log_exception_decorator(log_exception)
class Distributor(ctk.CTkToplevel):
    def __init__(self, folders_and_combinations=None, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.title(LANG.get("distributor"))
        self.grab_set()

        self.folders_and_combinations = folders_and_combinations

        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        self.dstr_frame = MyFrame(master=self)
        self.dstr_frame.grid()
        self.dstr_frame.columnconfigure(0, weight=1)
        self.dstr_frame.rowconfigure(0, weight=0)
        self.dstr_frame.rowconfigure((1, 2), weight=1)

        self.upper_pan = MyFrame(master=self.dstr_frame)
        self.upper_pan.grid(row=0, column=0)

        self.upper_pan.columnconfigure((0,1,2,3), weight=0)
        self.upper_pan.rowconfigure(0, weight=0)

        MyButton(master=self.upper_pan, text=LANG.get("go_back"),
                 command=self.destroy).grid(row=0, column=0)
        MyButton(master=self.upper_pan, text=LANG.get("create_files_list"),
                 command=self.create_files_list).grid(row=0, column=1)
        MyButton(master=self.upper_pan, text=LANG.get("create_folders_and_files_list"),
                 command=self.create_folders_with_files_list).grid(row=0, column=2)

        self.zoom_pan = MyFrame(master=self.dstr_frame)
        self.zoom_pan.grid(row=1, column=0)
        self.zoom_pan.columnconfigure(0, weight=1)
        self.zoom_pan.rowconfigure(0, weight=1)

        self.zoom_view = None

        self.detail_pan = MyFrame(master=self.dstr_frame)  # currently unused
        self.detail_pan.grid(row=2, column=0)
        self.detail_pan.columnconfigure(0, weight=1)
        self.detail_pan.rowconfigure(0, weight=1)

        self.detail_view = None
        self.files_creator = None

        self.tree_pan = MyFrame(master=self.dstr_frame)
        self.tree_pan.grid(row=2, column=0, columnspan=2)

        self.tree_pan.rowconfigure(0, weight=1)
        self.tree_pan.rowconfigure(1, weight=0)
        self.tree_pan.columnconfigure(0, weight=1)
        self.tree_pan.columnconfigure(1, weight=0)

        self.tree = ttk.Treeview(master=self.tree_pan, show="tree headings")
        self.tree.grid(row=0, column=0, sticky=tk.NSEW)

        scroll_y = ttk.Scrollbar(master=self.tree_pan, orient=tk.VERTICAL, command=self.tree.yview)
        scroll_y.grid(row=0, column=1, sticky=tk.NS)
        self.tree.configure(yscrollcommand=scroll_y.set)

        scroll_x = ttk.Scrollbar(master=self.tree_pan, orient=tk.HORIZONTAL, command=self.tree.xview)
        scroll_x.grid(row=1, column=0, sticky=tk.EW)
        self.tree.configure(xscrollcommand=scroll_x.set)

        self.tree.bind('<<TreeviewSelect>>', self.selecting_item)

        columns = ("name", "source", "path", "labels", "c_time")
        self.tree.configure(columns=columns)

        self.tree.heading("name", text=LANG.get("f_name"))
        self.tree.heading("source", text=LANG.get("source"))
        self.tree.heading("path", text=LANG.get("path"))
        self.tree.heading("labels", text=LANG.get("labels"))
        self.tree.heading("c_time", text=LANG.get("c_time"))

        self.project = self.master.brain.project

        self.populate_with_folders()

        my_logger.debug(f"Distributor initiated")

    def create_folders_with_files_list(self):
        d_path = filedialog.askdirectory(initialdir=os.getcwd(),
                                         title=LANG.get("create_folder"))

        if not d_path:
            return

        my_logger.debug(f"Distributor.create_folders_with_files_list: folder: {d_path}")

        d_path = pathlib.Path(d_path)
        if not d_path.exists():
            d_path.mkdir()

        report_path = filedialog.asksaveasfilename(initialdir=d_path,
                                                   title=LANG.get("create_report_file"),
                                                   defaultextension='.xlsx')

        my_logger.debug(f"Distributor.create_folders_with_files_list: report: {report_path}")

        if not report_path:
            return

        wb = Workbook()

        ws = wb.active
        ws.title = f"Report for {self.project['name']}"

        row = 1
        col = 1

        for folder, labels in self.folders_and_combinations.items():
            folder_path = d_path.joinpath(folder)
            if not folder_path.exists():
                folder_path.mkdir()

            ws.cell(row=row, column=col, value="Folder name")
            ws.cell(row=row, column=col + 1, value=folder)
            ws.cell(row=row, column=col + 2, value="Labels")
            ws.cell(row=row, column=col + 3, value=labels)

            row += 1

            ws.cell(row=row, column=col, value="File name")
            ws.cell(row=row, column=col + 1, value="Source")
            ws.cell(row=row, column=col + 2, value="Old Path")
            ws.cell(row=row, column=col + 3, value="New Path")
            ws.cell(row=row, column=col + 4, value="New Path Short")
            ws.cell(row=row, column=col + 5, value="Labels")
            ws.cell(row=row, column=col + 6, value="Creation time")

            row += 1

            labels = labels.split(" ___ ")
            unique_ids = set()

            for label in labels:
                for file_id, i in self.project["files"].items():
                    if label in i["labels"] and file_id not in unique_ids:
                        unique_ids.add(file_id)

                        old_path = self.master.brain.project_path.joinpath(i.get("path"))
                        new_path = folder_path.joinpath(i.get("f_name"))

                        new_path_short = f"{new_path.parent.parent.name}\\{new_path.parent.name}\\{new_path.name}"

                        new_path = str(new_path)

                        shutil.copy2(old_path, new_path)

                        ws.cell(row=row, column=col, value=i.get("f_name"))
                        ws.cell(row=row, column=col + 1, value=i.get("source"))
                        ws.cell(row=row, column=col + 2, value=str(old_path))
                        ws.cell(row=row, column=col + 3, value=str(new_path))
                        ws.cell(row=row, column=col + 4, value=new_path_short)
                        ws.cell(row=row, column=col + 5, value=" ___ ".join(i.get("labels")))
                        ws.cell(row=row, column=col + 6, value=dt.datetime.fromtimestamp(i.get("c_time")).strftime(
                                                                 "%Y-%m-%d %H:%M:%S"))
                        row += 1
            row += 1

        wb.save(report_path)

        messagebox.showinfo(master=self, title="", message=LANG.get("files_rep_saved"))

        my_logger.debug(f"Distributor.create_folders_with_files_list: folder and report saved")

    def create_files_list(self):
        report_path = filedialog.asksaveasfilename(master=self, initialdir=os.getcwd(), title=LANG.get("save_as"), defaultextension='.xlsx')

        if not report_path:
            return

        wb = Workbook()

        ws = wb.active
        ws.title = f"Report for {self.project['name']}"

        row = 1
        col = 1

        for folder, labels in self.folders_and_combinations.items():
            # root = f"{folder} $ {labels}"
            # root_num = id_num
            ws.cell(row=row, column=col, value="Folder name")
            ws.cell(row=row, column=col + 1, value=folder)
            ws.cell(row=row, column=col + 2, value="Labels")
            ws.cell(row=row, column=col + 3, value=labels)

            row += 1

            ws.cell(row=row, column=col, value="File name")
            ws.cell(row=row, column=col + 1, value="Source")
            ws.cell(row=row, column=col + 2, value="Path")
            ws.cell(row=row, column=col + 3, value="Labels")
            ws.cell(row=row, column=col + 4, value="Creation time")

            row += 1

            labels = labels.split(" ___ ")
            unique_ids = set()

            for label in labels:
                for file_id, i in self.project["files"].items():
                    if label in i["labels"] and file_id not in unique_ids:
                        unique_ids.add(file_id)
                        ws.cell(row=row, column=col, value=i.get("f_name"))
                        ws.cell(row=row, column=col + 1, value=i.get("source"))
                        ws.cell(row=row, column=col + 2, value=i.get("path"))
                        ws.cell(row=row, column=col + 3, value=" ___ ".join(i.get("labels")))
                        ws.cell(row=row, column=col + 4, value=dt.datetime.fromtimestamp(i.get("c_time")).strftime(
                                                                 "%Y-%m-%d %H:%M:%S"))
                        row += 1
            row += 1

        wb.save(report_path)

        messagebox.showinfo(master=self, title="", message=LANG.get("rep_saved"))

        my_logger.debug(f"Distributor.create_files_list: report saved")

    def selecting_item(self, event):
        # for child in self.detail_pan.winfo_children():  # currently unused
        #     child.destroy()

        for child in self.zoom_pan.winfo_children():
            child.destroy()

        item_index = self.tree.focus()
        item_obj = self.tree.item(item_index)
        item_vals = item_obj.get("values")

        try:
            path_to_screen = self.master.brain.project_path.joinpath(item_vals[2])
        except IndexError:
            return

        try:
            self.zoom_view = ZoomPan(master=self.zoom_pan, path_to_img=path_to_screen)
            self.zoom_view.grid()
            my_logger.debug(f"Distributor.selecting_item: {path_to_screen}")
        except tk.TclError:
            return
        except KeyError:
            return

        self.master.brain.collect_garbage()

    def populate_with_folders(self):

        my_logger.debug(f"Distributor.populate_with_folders")

        id_num = 0
        project = self.master.brain.project

        for folder, labels in self.folders_and_combinations.items():
            root = f"{folder} $ {labels}"
            root_num = id_num

            self.tree.insert("", index=root_num, iid=root_num, text=root, open=True)

            id_num += 1

            labels = labels.split(" ___ ")

            unique_ids = set()
            for label in labels:
                for file_id, i in project["files"].items():
                    if label in i["labels"] and file_id not in unique_ids:
                        unique_ids.add(file_id)
                        self.tree.insert(str(root_num),
                                         tk.END,
                                         values=(i.get("f_name"),
                                                 i.get("source"),
                                                 i.get("path"),
                                                 i.get("labels"),
                                                 dt.datetime.fromtimestamp(i.get("c_time")).strftime(
                                                                 "%Y-%m-%d %H:%M:%S")))


@log_exception_decorator(log_exception)
class SaveLoadSchema(ctk.CTkToplevel):
    def __init__(self, save=None, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.grab_set()
        self.save = save
        if self.save:
            self.title(LANG.get("save_schema"))
        else:
            self.title(LANG.get("load_schema"))

        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        self.pan = MyFrame(master=self)
        self.pan.grid()
        self.pan.columnconfigure(0, weight=1)
        self.pan.rowconfigure(0, weight=1)
        self.pan.rowconfigure((1, 2, 3, 4, 5), weight=0)

        self.lst_box = MyListbox(master=self.pan)
        self.lst_box.grid(row=0, column=0)

        self.schemas = self.master.master.master.brain.project["dist_schemas"]

        for schema in self.schemas:
            self.lst_box.insert(tk.END, schema)

        if self.save:

            MyLabel(master=self.pan, text=LANG.get("schema_name")).grid(row=1, column=0)
            self.ent_schema_name = MyEntry(self.pan)
            self.ent_schema_name.grid(row=2, column=0)

            MyButton(master=self.pan, text=LANG.get("save_schema"), command=self.save_schema).grid(row=3, column=0)
        else:
            MyLabel(master=self.pan, text=LANG.get("load_schema")).grid(row=1, column=0)
            MyButton(master=self.pan, text=LANG.get("load_schema"), command=self.load_schema).grid(row=2, column=0)

        MyButton(master=self.pan, text=LANG.get("del_schema"), command=self.del_schema).grid(row=4, column=0)
        MyButton(master=self.pan, text=LANG.get("go_back"), command=self.destroy).grid(row=5, column=0)

        my_logger.debug(f"SaveLoadSchema initiated as {self.save}")

    def save_schema(self):
        schema_name = self.ent_schema_name.get()
        if schema_name:
            my_logger.debug(f"SaveLoadSchema.save_schema {schema_name}")
            if schema_name in self.schemas:
                messagebox.showinfo(master=self, title=LANG.get("unable_to_save"), message=LANG.get("schema_exists"))
            else:
                labels_to_save = self.master.selection_listbox.get(0, tk.END)
                labels_combined_to_save = self.master.combined_listbox.get(0, tk.END)
                folders_and_combinations_to_save = self.master.folders_and_combinations

                self.schemas[schema_name] = {
                    "labels": labels_to_save,
                    "labels_combined": labels_combined_to_save,
                    "folders_and_combinations": folders_and_combinations_to_save
                }

                self.master.master.master.brain.project["dist_schemas"] = self.schemas
                self.master.master.master.brain.save_project()
                self.lst_box.insert(tk.END, schema_name)

                my_logger.debug(f"SaveLoadSchema.save_schema {schema_name} done")

    def load_schema(self):
        selected = self.lst_box.curselection()
        if selected:
            schema = self.lst_box.get(selected[0])

            labels_to_load = self.schemas[schema]["labels"]
            labels_combined_to_load = self.schemas[schema]["labels_combined"]
            folders_and_combinations_to_load = self.schemas[schema]["folders_and_combinations"]

            self.master.selection_listbox.delete(0, tk.END)
            self.master.selection_listbox.insert(tk.END, *labels_to_load)

            self.master.combined_listbox.delete(0, tk.END)
            self.master.combined_listbox.insert(tk.END, *labels_combined_to_load)

            self.master.folders_and_combinations = folders_and_combinations_to_load

            self.master.folder_listbox.delete(0, tk.END)
            self.master.folder_listbox.insert(0, *self.master.folders_and_combinations.keys())
            my_logger.debug(f"SaveLoadSchema.load_schema {schema} loaded")

    def del_schema(self):
        selected = self.lst_box.curselection()
        if selected:
            schema = self.lst_box.get(selected[0])
            del self.schemas[schema]
            self.master.master.master.brain.project["dist_schemas"] = self.schemas
            self.master.master.master.brain.save_project()

            self.lst_box.delete(selected[0])

            my_logger.debug(f"SaveLoadSchema.del_schema {schema} deleted")

