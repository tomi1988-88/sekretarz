import os
import json
import re
import pathlib
import shutil
import time
import copy
import gc
import tkinter as tk
import datetime as dt
from tkinter import (ttk,
                     messagebox,
                     filedialog,
                     simpledialog,
                     scrolledtext)
from __sekretarz_lang import LANG
from __zoomed_canvas import ZoomAdvanced
from openpyxl import Workbook


class MyFrame(ttk.Frame):
    def __init__(self, driver=None,
                 driver_frame=None,
                 driver_attr=None,
                 driver_var=None,
                 driver_tree_binded=None,
                 *args, **kwargs):
        super(MyFrame, self).__init__(*args, **kwargs)
        self.driver = driver
        self.driver_frame = driver_frame
        self.driver_attr = driver_attr
        self.driver_var = driver_var
        self.driver_tree_binded = driver_tree_binded

    def grid(self, *args, **kwargs):
        if kwargs.get("sticky"):
            super(MyFrame, self).grid(*args, **kwargs)
        else:
            super(MyFrame, self).grid(sticky=tk.NSEW, *args, **kwargs)


class ExtraField(ttk.Frame):
    def __init__(self, field_name=None, field_val=None, file_id=None, *args, **kwargs):
        super().__init__(**kwargs)

        self.field_name = field_name
        self.field_val = field_val
        self.file_id = file_id

        self.columnconfigure(0)
        self.columnconfigure(1)
        self.columnconfigure(2)

        self.lbl_field = ttk.Label(master=self, text=self.field_name)
        self.lbl_field.grid(column=0, row=0)
        self.ent_field = ttk.Entry(master=self, width=10)
        self.ent_field.grid(column=1, row=0)
        self.ent_field.insert(tk.END, self.field_val)
        self.btn_field = ttk.Button(master=self, text=LANG.get("update"), command=self.update)
        self.btn_field.grid(column=2, row=0)

        self.project = self.nametowidget(".").load_project()

    def update(self):
        self.field_val = self.ent_field.get()

        self.project["files"][self.file_id]["extra_fields"][self.lbl_field["text"]] = self.field_val

        self.nametowidget(".").project = self.project
        self.nametowidget(".").save_project()

    def get_values(self):
        return self.file_id, self.field_name, self.field_val


class NewProWindow(tk.Toplevel):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.geometry("600x100")
        self.title(f'{LANG.get("new_pro")}')
        self.grab_set()

        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        self.main_frame = MyFrame(master=self)
        self.main_frame.grid()
        self.main_frame.columnconfigure(0, weight=1)
        self.main_frame.columnconfigure(1, weight=1)
        self.main_frame.rowconfigure(0, weight=1)
        self.main_frame.rowconfigure(1, weight=1)
        self.main_frame.rowconfigure(2, weight=1)

        ttk.Label(master=self.main_frame, text=LANG.get("proj_name")).grid(row=0, column=0)

        self.ent_pro_name = ttk.Entry(master=self.main_frame)
        self.ent_pro_name.grid(row=0, column=1, sticky=tk.EW, padx=10)

        ttk.Label(master=self.main_frame, text=LANG.get("proj_local")).grid(row=1, column=0)

        self.curr_dir = os.getcwd()
        ttk.Label(master=self.main_frame, text=self.curr_dir).grid(row=1, column=1)

        ttk.Button(master=self.main_frame, text=LANG.get("go_back"), command=self.destroy).grid(row=2, column=0)

        ttk.Button(
            master=self.main_frame,
            text=LANG.get("new_pro"),
            command=lambda x=self.ent_pro_name.get: self.create_project(x())
        ).grid(row=2, column=1)

        self.pat_format = re.compile(r"[<>:\"/\\|?*]")

    def create_project(self, project_name):
        if self.pat_format.search(project_name) or project_name == "":
            messagebox.showerror(master=self, title=f'{LANG.get("error")}', message=f'{LANG.get("wrong_dir_name")}')
            return

        path = pathlib.Path(self.curr_dir, project_name)

        if path.exists():
            messagebox.showerror(master=self, title=f'{LANG.get("error")}', message=f'{LANG.get("dir_exists")}')
            return

        path.mkdir()

        base_proj = {
            "name": project_name,
            "main_dir": str(path),
            "files": {},
            "labels": [],
            "last_id": 0,
            "links": [],
            "binds": {},
            "last_id_binds": 0,
            "dist_schemas": {}
        }

        with open(pathlib.Path(path, "base.json"), mode="w") as file:
            json.dump(base_proj, file, indent=4)

        self.master.project = base_proj
        self.master.base_project_view()
        self.destroy()


class BaseProjectView(MyFrame):
    def __init__(self, *args, **kwargs):
        super(BaseProjectView, self).__init__(*args, **kwargs)

        self.main_window = self.nametowidget(".")
        self.main_window.title(f'{LANG.get("title")}: {self.main_window.project["main_dir"]}')

        # self.columnconfigure(0, weight=1, minsize=700)
        # self.columnconfigure(1, weight=1, minsize=1000)

        self.columnconfigure(0, weight=1, minsize=500)
        self.columnconfigure(1, weight=1, minsize=800)

        self.rowconfigure(0, weight=0)
        self.rowconfigure(1, weight=1)
        self.rowconfigure(2, weight=1)

        self.menu_bar = MyFrame(master=self)
        self.menu_bar.grid(row=0, column=0, sticky=tk.W)

        self.project = self.main_window.load_project()

        ttk.Button(master=self.menu_bar, text=LANG.get("add_folder"),
                   command=self.add_folder).pack(side=tk.LEFT, padx=5, pady=2)

        ttk.Button(master=self.menu_bar, text=LANG.get("manage_links"),
                   command=self.source_manager).pack(side=tk.LEFT, padx=5, pady=2)

        # ttk.Button(master=self.menu_bar, text=LANG.get("list_files"), command=self.list_files).pack(side=tk.LEFT) #todo
        #
        ttk.Button(master=self.menu_bar, text=LANG.get("labels"), command=self.label_manager).pack(side=tk.LEFT)
        #
        # ttk.Button(master=self.menu_bar, text=LANG.get("del_pro"), ).pack(side=tk.LEFT) #todo
        #
        ttk.Button(master=self.menu_bar, text=LANG.get("filter_labels"),
                   command=self.filter_manager).pack(side=tk.LEFT)
        # ttk.Button(master=self.menu_bar, text=LANG.get("bind_manager"), #todo
        #            command=self.binding_manager).pack(side=tk.LEFT)
        ttk.Button(master=self.menu_bar, text=LANG.get("dist_manager"),
                   command=self.distribute_labels_to_dirs).pack(side=tk.LEFT)

        ttk.Label(master=self.menu_bar, text=f'{LANG.get("proj_name")} {self.project.get("name")}').pack(padx=15,
                                                                                                         pady=10)

        self.left_pan = MyFrame(master=self)
        self.left_pan.grid(row=1, column=0, rowspan=2, )

        self.left_pan.columnconfigure(0, weight=1)
        self.left_pan.rowconfigure(0, weight=1)
        self.left_pan.rowconfigure(1, weight=1)

        self.tree_view = TreePan(master=self.left_pan, driver=self)
        self.tree_view.grid(row=0, column=0, )

        self.detail_pan = MyFrame(master=self.left_pan, driver=self)
        self.detail_pan.grid(row=1, column=0, )
        self.detail_pan.columnconfigure(0, weight=1)
        self.detail_pan.rowconfigure(0, weight=1)

        self.right_pan = MyFrame(master=self, driver=self)
        self.right_pan.grid(row=0, column=1, rowspan=3, )

        self.right_pan.columnconfigure(0, weight=1)
        self.right_pan.rowconfigure(0, weight=1)

        self.detail_view = None
        self.zoom_view = None

        self.pat_formats = re.compile(r"(.png$|.jpg$|.jpeg$)", flags=re.IGNORECASE)
        self.src_manager = None
        self.lbl_manager = None
        self.flt_manager = None
        self.bin_manager = None
        self.dist_manager = None

        if self.project["files"]:
            self.tree_view.populate()
        else:
            ttk.Label(master=self.detail_pan, text=LANG.get("no_files")).grid()

    def _add_folder__ignore_patterns(self, path, names):
        return [f for f in names if pathlib.Path(path, f).is_file() and not self.pat_formats.search(f)]

    def add_folder(self):
        d_path = filedialog.askdirectory(initialdir=os.getcwd(), title=LANG.get("add_folder"))
        if not d_path:
            return

        d_path = pathlib.Path(d_path)
        main_dir = self.project["main_dir"]

        target_dir = pathlib.Path(main_dir, d_path.stem)

        if target_dir.exists():
            target_dir = pathlib.Path(main_dir, d_path.stem + str(int(time.time())))

        shutil.copytree(d_path, target_dir, ignore=self._add_folder__ignore_patterns)

        pro_counter = int(self.project["last_id"])

        for (dpath, dirs, files) in os.walk(target_dir):
            for file in files:
                path = pathlib.Path(dpath, file)
                pro_counter += 1
                self.project["files"][f"{pro_counter} - {path.name}"] = {
                    "id": pro_counter,
                    "source": "",
                    "f_name": path.name,
                    "path": str(path),
                    "labels": [],
                    "comment": "",
                    "extra_fields": {},
                    "c_time": path.stat().st_mtime,
                    "binds": [],
                }

        files_added = pro_counter - self.project["last_id"]

        self.project["last_id"] = pro_counter

        print(self.project)

        self.nametowidget(".").project = self.project
        self.nametowidget(".").save_project()
        self.reset_tree()

        if files_added > 0:
            messagebox.showinfo(master=self, title=LANG.get("add_folder"), message=f'{LANG.get("update_copy_finish")}{files_added}')
            self.reset_tree()
        else:
            messagebox.showerror(master=self, title=LANG.get("add_folder"), message=LANG.get("no_files_added"))

    def reset_tree(self):
        self.project = self.nametowidget(".").load_project()
        self.tree_view.populate()
        gc.disable()
        gc.collect()
        gc.enable()

    def source_manager(self):
        self.src_manager = SourceManager(master=self)

    def label_manager(self):
        self.lbl_manager = LabelManager(master=self)

    def filter_manager(self):
        self.flt_manager = FilterManager()

    def distribute_labels_to_dirs(self):
        self.dist_manager = DistManager()

    def manage_all_fields(self):
        # todo
        ...

    def binding_manager(self):
        # self.bin_manager = BindingManager()
        # todo
        # relacje: vertical (parent, child,) horizontal (sibling)
        # add "binds" to the project
        # binds structure: {bind_id: {vertical: (parent, child)} OR {horizontal: [sibling, sibling, sibling]}}
        # add "bind_id" to project["files"][file_id]
        # view:
        # cols: 3
        # rows: 2
        # col 0 row 0: zoom_view_0
        # col 1 row 0: zoom_view_1
        # col 2 row 0: zoom_view_2

        # col 0 row 1: tree: binds of all files 0
        # col 1, 2 row 1 - colspan=2: down_pan
        # down_pan: cols: 3, row: 1
        # down_pan 0, 0 = tree: lst of files 1
        # down_pan 1, 0 = actions
        # down_pan 2, 0 = tree: lst of files 2

        # actions:
        # bind as parent to - ignore if bind exists
        # bind as child to - ignore if bind exists
        # bind as sibling to
        # unbind as parent/child
        # unbind as sibling
        # show all binds
        # go back

        # alternatywa:
        # tutaj spierdolone - trzeba inaczej zrobić: przemodelować cały Binding Manager
        # w ten sposób, że po lewej jest TreeBinded po prawej panel z listą
        # zaznaczam na liście obiekt i do niego akcje:
        # utowórz root
        # add to selected root as parent or child
        # unbind item selected in TreeBinded etc
        ...


class DistManager(tk.Toplevel):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.title(LANG.get("dist_manager"))
        # width = self.winfo_screenwidth()
        # height = self.winfo_screenheight()
        # self.geometry(f"{width}x{height}")
        self.grab_set()
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        self.dist_frame = MyFrame(master=self)
        self.dist_frame.grid()

        self.dist_frame.columnconfigure((0, 1, 2), weight=1)
        self.dist_frame.rowconfigure(0)
        self.dist_frame.rowconfigure(1)
        self.dist_frame.rowconfigure(2, weight=1)
        self.dist_frame.rowconfigure(3)

        self.upper_pan = MyFrame(master=self.dist_frame)
        self.upper_pan.grid(row=0, column=0)

        ttk.Button(master=self.upper_pan, text="Load Schema", command=self.load_schema).grid()
        ttk.Button(master=self.upper_pan, text="Save Schema", command=self.save_schema).grid()

        ttk.Label(master=self.dist_frame, text="Select Labels").grid(row=1, column=0)
        ttk.Label(master=self.dist_frame, text="Combined Labels").grid(row=1, column=1)
        ttk.Label(master=self.dist_frame, text="Folders").grid(row=1, column=2)

        self.left_pan = MyFrame(master=self.dist_frame)
        self.left_pan.grid(row=2, column=0)
        self.left_pan.rowconfigure(0, weight=1)
        self.left_pan.columnconfigure(0, weight=1)

        self.label_lst = tk.Listbox(master=self.left_pan, selectmode=tk.MULTIPLE)
        self.label_lst.grid(sticky=tk.NSEW)

        self.center_pan = MyFrame(master=self.dist_frame,)
        self.center_pan.grid(row=2, column=1)
        self.center_pan.rowconfigure(0, weight=1)
        self.center_pan.columnconfigure(0, weight=1)

        self.combined_lst = tk.Listbox(master=self.center_pan, exportselection=False)
        self.combined_lst.grid(sticky=tk.NSEW)

        self.right_pan = MyFrame(master=self.dist_frame)
        self.right_pan.grid(row=2, column=2)
        self.right_pan.rowconfigure(0, weight=1)
        self.right_pan.columnconfigure(0, weight=1)

        self.folders_lst = tk.Listbox(master=self.right_pan, exportselection=False)
        self.folders_lst.grid(sticky=tk.NSEW)

        self.project = self.nametowidget(".").load_project()
        self.labels = self.project["labels"]

        for l in self.labels:
            self.label_lst.insert(tk.END, l)

        self.bottom_left = MyFrame(master=self.dist_frame)
        self.bottom_left.grid(row=3, column=0)
        self.bottom_left.rowconfigure((0,1,2), weight=0)
        self.bottom_left.columnconfigure(0, weight=1)
        ttk.Label(master=self.bottom_left, text="Select Label(s) and push them to Combined Labels").grid()
        ttk.Button(master=self.bottom_left, text="Push selected Labels", command=self.push_selected_label).grid() #todo

        self.bottom_center = MyFrame(master=self.dist_frame)
        self.bottom_center.grid(row=3, column=1)
        self.bottom_center.rowconfigure((0,1,2,3), weight=0)
        self.bottom_center.columnconfigure(0, weight=1)
        ttk.Label(master=self.bottom_center, text="Check if combination is ok and push them to Folders").grid()
        ttk.Button(master=self.bottom_center, text="Push all combined Labels to Folders", command=self.push_all_combined_labels).grid()
        ttk.Button(master=self.bottom_center, text="Push combined Label to Folders", command=self.push_combined_label).grid()
        ttk.Button(master=self.bottom_center, text="Delete combination", command=self.del_combined).grid()

        self.bottom_right = MyFrame(master=self.dist_frame)
        self.bottom_right.grid(row=3, column=2)
        self.bottom_right.rowconfigure((0,1,2,3), weight=0)
        self.bottom_right.columnconfigure(0, weight=1)
        ttk.Label(master=self.bottom_right, text="Rename Folders or keep default names. If all folders are ready start Distribution").grid()
        ttk.Button(master=self.bottom_right, text="Rename Folder", command=self.rename_folder).grid()
        ttk.Button(master=self.bottom_right, text="Prepare Distribution", command=self.prep_distribution).grid()
        ttk.Button(master=self.bottom_right, text="Delete folder", command=self.del_folder).grid()

        self.folders_and_combinations = dict()

        self.new_folder_name = None
        self.distributor = None
        self.save_load = None

    def save_schema(self):
        self.save_load = SaveLoadSchema(master=self, save=True)

    def load_schema(self):
        self.save_load = SaveLoadSchema(master=self, save=False)

    def prep_distribution(self):
        if self.folders_and_combinations:
            self.distributor = Distributor(self.folders_and_combinations)
        else:
            messagebox.showinfo(master=self, title=LANG.get("dist_manager"), message="Prepare Folders first")

    def rename_folder(self):
        self.new_folder_name = None
        selected = self.folders_lst.curselection()
        if selected:
            item_id = selected[0]
            old_name = self.folders_lst.get(item_id)

            self.new_folder_name = simpledialog.askstring(title=f"Rename {old_name}", prompt=f"Rename {old_name}")

            if self.new_folder_name:
                self.folders_and_combinations[self.new_folder_name] = self.folders_and_combinations[old_name]
                del self.folders_and_combinations[old_name]
                print(self.folders_and_combinations)

                self.folders_lst.delete(item_id)
                self.folders_lst.insert(item_id, self.new_folder_name)

    def push_all_combined_labels(self):
        all_comb_labels = self.combined_lst.get(0, tk.END)
        for comb_label in all_comb_labels:
            key_folder = comb_label

            if key_folder in self.folders_and_combinations:
                return

            self.folders_and_combinations[key_folder] = key_folder
            self.folders_lst.insert(tk.END, key_folder)

    def push_selected_label(self):
        selected = self.label_lst.curselection()
        if selected:
            selected = [self.label_lst.get(i) for i in selected]
            selected = " ___ ".join(selected)

            self.combined_lst.insert(tk.END, selected)

    def push_combined_label(self):
        selected = self.combined_lst.curselection()
        if selected:
            key_folder = self.combined_lst.get(selected[0])

            if key_folder in self.folders_and_combinations:
                return

            self.folders_and_combinations[key_folder] = key_folder
            self.folders_lst.insert(tk.END, key_folder)

    def del_combined(self):
        selected = self.combined_lst.curselection()
        if selected:
            item_in_folder = self.combined_lst.get(selected[0])

            for key, val in self.folders_and_combinations.items():
                if item_in_folder == val:
                    res = messagebox.askyesno(master=self, title=LANG.get("dist_manager"), message=f"This operation will affect Folder: {key}")
                    if res:
                        self.combined_lst.delete(selected[0])
                        self.folders_lst.delete(0, tk.END)
                        del self.folders_and_combinations[key]

                        for key, val in self.folders_and_combinations.items():
                            self.folders_lst.insert(tk.END, val)
                        break
                    break
            self.combined_lst.delete(selected[0])

    def del_folder(self):
        selected = self.folders_lst.curselection()
        if selected:
            del self.folders_and_combinations[self.folders_lst.get(selected[0])]
            self.folders_lst.delete(selected[0])


class SaveLoadSchema(tk.Toplevel):
    def __init__(self, save=None, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.grab_set()
        self.save = save
        if self.save:
            self.title("Save Schema")
        else:
            self.title("Load Schema")

        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        self.pan = MyFrame(master=self)
        self.pan.grid()
        self.pan.columnconfigure(0, weight=1)
        self.pan.rowconfigure(0, weight=1)
        self.pan.rowconfigure((1, 2, 3, 4, 5), weight=0)

        self.lst_box = tk.Listbox(master=self.pan)
        self.lst_box.grid(row=0, column=0)

        self.project = self.nametowidget(".").load_project()

        if "dist_schemas" in self.project:
            self.schemas = self.project["dist_schemas"]
        else:
            self.project["dist_schemas"] = dict()
            self.schemas = self.project["dist_schemas"]

        for schema in self.schemas:
            self.lst_box.insert(tk.END, schema)

        if self.save:

            ttk.Label(master=self.pan, text="Schema name:").grid(row=1, column=0)
            self.ent_schema_name = ttk.Entry(self.pan)
            self.ent_schema_name.grid(row=2, column=0)

            ttk.Button(master=self.pan, text="Save", command=self.save_schema).grid(row=3, column=0)
        else:
            ttk.Label(master=self.pan, text="Load Schema").grid(row=1, column=0)
            ttk.Button(master=self.pan, text="Load", command=self.load_schema).grid(row=2, column=0)

        ttk.Button(master=self.pan, text="Delete", command=self.del_schema).grid(row=4, column=0)
        ttk.Button(master=self.pan, text=LANG.get("go_back"), command=self.destroy).grid(row=5, column=0)

    def save_schema(self):
        schema_name = self.ent_schema_name.get()
        if schema_name:
            if schema_name in self.schemas:
                messagebox.showinfo(marter=self, title="Unable to save", message="Schema with this name already exists")
            else:
                labels_to_save = self.master.label_lst.get(0, tk.END)
                labels_combined_to_save = self.master.combined_lst.get(0, tk.END)
                folders_and_combinations_to_save = self.master.folders_and_combinations

                self.project["dist_schemas"][schema_name] = {
                    "labels": labels_to_save,
                    "labels_combined": labels_combined_to_save,
                    "folders_and_combinations": folders_and_combinations_to_save
                }

                self.schemas = self.project["dist_schemas"]
                self.nametowidget(".").save_project()
                self.lst_box.insert(tk.END, schema_name)

    def load_schema(self):
        selected = self.lst_box.curselection()
        if selected:
            schema = self.lst_box.get(selected[0])

            labels_to_load = self.schemas[schema]["labels"]
            labels_combined_to_load = self.schemas[schema]["labels_combined"]
            folders_and_combinations_to_load = self.schemas[schema]["folders_and_combinations"]

            self.master.label_lst.delete(0, tk.END)
            self.master.label_lst.insert(tk.END, *labels_to_load)

            self.master.combined_lst.delete(0, tk.END)
            self.master.combined_lst.insert(tk.END, *labels_combined_to_load)

            self.master.folders_and_combinations = folders_and_combinations_to_load

            self.master.folders_lst.delete(0, tk.END)
            self.master.folders_lst.insert(0, *self.master.folders_and_combinations.keys())

    def del_schema(self):
        selected = self.lst_box.curselection()
        if selected:
            schema = self.lst_box.get(selected[0])
            del self.schemas[schema]
            self.nametowidget(".").save_project()
            self.lst_box.delete(selected[0])


class Distributor(tk.Toplevel):
    def __init__(self, folders_and_combinations=None, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.title("Distributor")
        self.grab_set()

        self.folders_and_combinations = folders_and_combinations

        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        self.dstr_frame = MyFrame(master=self)
        self.dstr_frame.grid()
        self.dstr_frame.columnconfigure((0, 1), weight=1)
        self.dstr_frame.rowconfigure(0)
        self.dstr_frame.rowconfigure((1, 2), weight=1)

        self.upper_pan = MyFrame(master=self.dstr_frame)
        self.upper_pan.grid(row=0, column=0, columnspan=1)

        self.upper_pan.columnconfigure((0,1,2,3), weight=0)
        self.upper_pan.rowconfigure(0)

        ttk.Button(master=self.upper_pan, text=LANG.get("go_back"), command=self.destroy).grid(row=0, column=0)
        ttk.Button(master=self.upper_pan, text="Create Files list", command=self.create_files_list).grid(row=0, column=1) # todo
        ttk.Button(master=self.upper_pan, text="Create Folders and Files list", command=self.create_folders_with_files_list).grid(row=0, column=2) # todo

        self.zoom_pan = MyFrame(master=self.dstr_frame)
        self.zoom_pan.grid(row=1, column=0)
        self.zoom_pan.columnconfigure(0, weight=1)
        self.zoom_pan.rowconfigure(0, weight=1)

        self.zoom_view = None

        self.detail_pan = MyFrame(master=self.dstr_frame)
        self.detail_pan = MyFrame(master=self.dstr_frame)
        self.detail_pan.grid(row=1, column=1)
        self.detail_pan.columnconfigure(0, weight=1)
        self.detail_pan.rowconfigure(0, weight=1)

        self.detail_view = None
        self.files_creator = None


        self.tree_pan = MyFrame(master=self.dstr_frame)
        self.tree_pan.grid(row=2, column=0, columnspan=2)

        self.tree_pan.rowconfigure(0, weight=1)
        self.tree_pan.rowconfigure(1, weight=0)
        self.tree_pan.columnconfigure(0, weight=1)
        self.tree_pan.columnconfigure(1, weight=0)

        self.tree = ttk.Treeview(master=self.tree_pan, show="tree headings")
        self.tree.grid(row=0, column=0, sticky=tk.NSEW)

        scroll_y = ttk.Scrollbar(master=self.tree_pan, orient=tk.VERTICAL, command=self.tree.yview)
        scroll_y.grid(row=0, column=1, sticky=tk.NS)
        self.tree.configure(yscrollcommand=scroll_y.set)

        scroll_x = ttk.Scrollbar(master=self.tree_pan, orient=tk.HORIZONTAL, command=self.tree.xview)
        scroll_x.grid(row=1, column=0, sticky=tk.EW)
        self.tree.configure(xscrollcommand=scroll_x.set)

        self.tree.bind('<<TreeviewSelect>>', self.selecting_item)

        columns = ("id", "source", "path", "labels", "c_time")
        self.tree.configure(columns=columns)

        self.tree.heading("id", text=LANG.get("id"))
        self.tree.heading("source", text=LANG.get("source"))
        self.tree.heading("path", text=LANG.get("path"))
        self.tree.heading("labels", text=LANG.get("labels"))
        self.tree.heading("c_time", text=LANG.get("c_time"))

        self.project = self.nametowidget(".").load_project()

        self.populate_with_folders()

    def create_folders_with_files_list(self):
        d_path = filedialog.askdirectory(master=self, initialdir=os.getcwd(), title="Create folder")

        if not d_path:
            return

        d_path = pathlib.Path(d_path)
        if not d_path.exists():
            d_path.mkdir()

        raport_path = filedialog.asksaveasfilename(master=self, initialdir=d_path, title="Create Raport file - xlsx extension is fixed", defaultextension='.xlsx')

        if not raport_path:
            return

        wb = Workbook()

        ws = wb.active
        ws.title = f"Raport for {self.project['name']}"

        row = 1
        col = 1

        for folder, labels in self.folders_and_combinations.items():
            folder_path = d_path.joinpath(folder)
            folder_path.mkdir()

            ws.cell(row=row, column=col, value="Folder name")
            ws.cell(row=row, column=col + 1, value=folder)
            ws.cell(row=row, column=col + 2, value="Labels")
            ws.cell(row=row, column=col + 3, value=labels)

            row += 1

            ws.cell(row=row, column=col, value="File name")
            ws.cell(row=row, column=col + 1, value="Source")
            ws.cell(row=row, column=col + 2, value="Old Path")
            ws.cell(row=row, column=col + 3, value="New Path")
            ws.cell(row=row, column=col + 4, value="New Path Short")
            ws.cell(row=row, column=col + 5, value="Labels")
            ws.cell(row=row, column=col + 6, value="Creation time")

            row += 1

            labels = labels.split(" ___ ")
            unique_ids = set()

            for label in labels:
                for file_id, i in self.project["files"].items():
                    if label in i["labels"] and file_id not in unique_ids:
                        unique_ids.add(file_id)

                        old_path = i.get("path")
                        new_path = folder_path.joinpath(file_id)

                        new_path_short = f"{new_path.parent.parent.name}\\{new_path.parent.name}\\{new_path.name}"

                        new_path = str(new_path)

                        shutil.copy2(old_path, new_path)

                        ws.cell(row=row, column=col, value=file_id)
                        ws.cell(row=row, column=col + 1, value=i.get("source"))
                        ws.cell(row=row, column=col + 2, value=old_path)
                        ws.cell(row=row, column=col + 3, value=new_path)
                        ws.cell(row=row, column=col + 4, value=new_path_short)
                        ws.cell(row=row, column=col + 5, value=" ___ ".join(i.get("labels")))
                        ws.cell(row=row, column=col + 6, value=dt.datetime.fromtimestamp(i.get("c_time")).strftime(
                                                                 "%Y-%m-%d %H:%M:%S"))
                        row += 1
            row += 1

        wb.save(raport_path)

        messagebox.showinfo(master=self, title="", message="Files and Raport saved!")

    def create_files_list(self):
        raport_path = filedialog.asksaveasfilename(master=self, initialdir=os.getcwd(), title="Save as", defaultextension='.xlsx')

        if not raport_path:
            return

        wb = Workbook()

        ws = wb.active
        ws.title = f"Raport for {self.project['name']}"

        row = 1
        col = 1

        for folder, labels in self.folders_and_combinations.items():
            # root = f"{folder} $ {labels}"
            # root_num = id_num
            ws.cell(row=row, column=col, value="Folder name")
            ws.cell(row=row, column=col + 1, value=folder)
            ws.cell(row=row, column=col + 2, value="Labels")
            ws.cell(row=row, column=col + 3, value=labels)

            row += 1

            ws.cell(row=row, column=col, value="File name")
            ws.cell(row=row, column=col + 1, value="Source")
            ws.cell(row=row, column=col + 2, value="Path")
            ws.cell(row=row, column=col + 3, value="Labels")
            ws.cell(row=row, column=col + 4, value="Creation time")

            row += 1

            labels = labels.split(" ___ ")
            unique_ids = set()

            for label in labels:
                for file_id, i in self.project["files"].items():
                    if label in i["labels"] and file_id not in unique_ids:
                        unique_ids.add(file_id)
                        ws.cell(row=row, column=col, value=file_id)
                        ws.cell(row=row, column=col + 1, value=i.get("source"))
                        ws.cell(row=row, column=col + 2, value=i.get("path"))
                        ws.cell(row=row, column=col + 3, value=" ___ ".join(i.get("labels")))
                        ws.cell(row=row, column=col + 4, value=dt.datetime.fromtimestamp(i.get("c_time")).strftime(
                                                                 "%Y-%m-%d %H:%M:%S"))
                        row += 1
            row += 1

        wb.save(raport_path)

        messagebox.showinfo(master=self, title="", message="Raport saved!")

    def selecting_item(self, event):
        for child in self.detail_pan.winfo_children():
            child.destroy()

        for child in self.zoom_pan.winfo_children():
            child.destroy()

        item_index = self.tree.focus()
        item_obj = self.tree.item(item_index)
        item_vals = item_obj.get("values")

        try:
            # file_id = item_vals[0]
            path_to_screen = item_vals[2]
        except IndexError:
            return

        try:
            # self.detail_view = DetailPan(master=self.detail_pan, driver=self,
            #                                     item_index=item_index, file_id=file_id)

            # comment = self.project["files"][file_id]["comment"]

            self.zoom_view = ZoomPan(master=self.zoom_pan, path=path_to_screen)

            gc.disable()
            gc.collect()
            gc.enable()

        except tk.TclError:
            return
        except KeyError:
            return

    def populate_with_folders(self):

        id_num = 0
        project = self.nametowidget(".").load_project()

        for folder, labels in self.folders_and_combinations.items():
            root = f"{folder} $ {labels}"
            root_num = id_num

            self.tree.insert("", index=root_num, iid=root_num, text=root, open=True)

            id_num += 1

            labels = labels.split(" ___ ")

            unique_ids = set()
            for label in labels:
                for file_id, i in project["files"].items():
                    if label in i["labels"] and file_id not in unique_ids:
                        unique_ids.add(file_id)
                        self.tree.insert(str(root_num), tk.END, values=(file_id, i.get("source"), i.get("path"), i.get("labels"),
                                                             dt.datetime.fromtimestamp(i.get("c_time")).strftime(
                                                                 "%Y-%m-%d %H:%M:%S")))


class BindingManager(tk.Toplevel):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.title(LANG.get("bind_manager"))
        width = self.winfo_screenwidth()
        height = self.winfo_screenheight()
        self.geometry(f"{width}x{height}")
        self.grab_set()
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        self.bind_frame = MyFrame(master=self)
        self.bind_frame.grid()

        self.bind_frame.columnconfigure((0, 1, 2), weight=1)
        self.bind_frame.rowconfigure((0, 1), weight=1)

        # style_1 = ttk.Style()
        # style_1.configure("_1.TFrame", background="black")
        #
        # style_1.configure("_2.TFrame", background="yellow")

        self.zoom_frame_0 = MyFrame(master=self.bind_frame)
        self.zoom_frame_0.grid(column=0, row=0)
        self.zoom_frame_0.columnconfigure(0, weight=1)
        self.zoom_frame_0.rowconfigure(0, weight=1)
        self.zoom_view_0 = ZoomPan(master=self.zoom_frame_0)

        self.zoom_frame_1 = MyFrame(master=self.bind_frame)
        self.zoom_frame_1.grid(column=1, row=0)
        self.zoom_frame_1.columnconfigure(0, weight=1)
        self.zoom_frame_1.rowconfigure(0, weight=1)
        self.zoom_view_1 = ZoomPan(master=self.zoom_frame_1)

        self.zoom_frame_2 = MyFrame(master=self.bind_frame)
        self.zoom_frame_2.grid(column=2, row=0)
        self.zoom_frame_2.columnconfigure(0, weight=1)
        self.zoom_frame_2.rowconfigure(0, weight=1)
        self.zoom_view_2 = ZoomPan(master=self.zoom_frame_2)

        self.tree_binded = TreeBinded(master=self.bind_frame, driver_frame=self.zoom_frame_0, driver_attr=self.zoom_view_0)
        self.tree_binded.grid(column=0, row=1)
        # self.tree_binded.populate()
        self.file_selected = [None]
        self.file_to_bind = [None]

        self.bottom_pan = MyFrame(master=self.bind_frame)
        self.bottom_pan.grid(column=1, row=1, columnspan=2)
        self.bottom_pan.columnconfigure((0, 1, 2), weight=1)
        self.bottom_pan.rowconfigure(0, weight=1)

        self.bottom_pan_left = MyFrame(master=self.bottom_pan)
        self.bottom_pan_left.grid(column=0, row=0)
        self.bottom_pan_left.columnconfigure(0, weight=1)
        self.bottom_pan_left.rowconfigure(0, weight=1)
        self.bottom_pan_centre = MyFrame(master=self.bottom_pan, style="_1.TFrame")
        self.bottom_pan_centre.grid(column=1, row=0)
        self.bottom_pan_centre.columnconfigure(0, weight=1)
        self.bottom_pan_centre.rowconfigure((0, 1, 2, 3, 4, 5, 6), weight=1)
        self.bottom_pan_right = MyFrame(master=self.bottom_pan, style="_1.TFrame")
        self.bottom_pan_right.grid(column=2, row=0)
        self.bottom_pan_right.columnconfigure(0, weight=1)
        self.bottom_pan_right.rowconfigure(0, weight=1)

        self.tree_choice = TreeChoiceOrToBind(master=self.bottom_pan_left,
                                              driver_frame=self.zoom_frame_1,
                                              driver_attr=self.zoom_view_1,
                                              driver_var=self.file_selected,
                                              driver_tree_binded=self.tree_binded)
        self.tree_choice.grid()
        self.tree_choice.populate()

        ttk.Button(master=self.bottom_pan_centre, text=LANG.get("show_binds"), command=self.show_binds).grid()
        ttk.Button(master=self.bottom_pan_centre, text=LANG.get("bind_parent"), command=self.bind_parent).grid()
        ttk.Button(master=self.bottom_pan_centre, text=LANG.get("bind_child"), command=self.bind_child).grid()
        ttk.Button(master=self.bottom_pan_centre, text=LANG.get("bind_sibling"), command=self.bind_sibling).grid()
        ttk.Button(master=self.bottom_pan_centre, text=LANG.get("unbind_parent_child"), command=self.unbind_parent_child).grid()
        ttk.Button(master=self.bottom_pan_centre, text=LANG.get("unbind_sibling"), command=self.unbind_parent_child).grid()
        ttk.Button(master=self.bottom_pan_centre, text=LANG.get("show_all_binds"), command=self.show_all_binds).grid()
        ttk.Button(master=self.bottom_pan_centre, text=LANG.get("go_back"), command=self.destroy).grid()

        self.tree_to_bind = TreeChoiceOrToBind(master=self.bottom_pan_right,
                                               driver_frame=self.zoom_frame_2,
                                               driver_attr=self.zoom_view_2,
                                               driver_var=self.file_to_bind,
                                               driver_tree_binded=self.tree_binded)
        self.tree_to_bind.grid()
        self.tree_to_bind.populate()

        self.project = None

    def show_binds(self):
        if self.file_selected[0]:
            self.tree_binded.show_binds(file_id=self.file_selected[0])
        else:
            messagebox.showinfo(master=self, title=LANG.get("bind_manager"), message=LANG.get("select_file_item"))

    def bind_parent(self):

        print(self.file_selected)
        print(self.file_to_bind)
        file_selected = self.file_selected[0]
        file_to_bind = self.file_to_bind[0]

        if file_selected and file_to_bind:
            self.project = self.nametowidget(".").load_project()

            self.project["last_id_binds"] += 1
            bind_num = self.project["last_id_binds"]

            self.project["binds"][bind_num] = {"vertical": (file_selected, file_to_bind)}

            self.project["files"][file_selected]["binds"].append(bind_num)
            self.project["files"][file_to_bind]["binds"].append(bind_num)


            self.nametowidget(".").project = self.project
            self.nametowidget(".").save_project()

            self.tree_binded.show_binds(file_selected)

        else:
             messagebox.showinfo(master=self, title=LANG.get("bind_manager"), message=LANG.get("select_files_to_bind"))
    def bind_child(self):
        ...
    def bind_sibling(self):
        ...
    def unbind_parent_child(self):
        ...
    def show_all_binds(self):
        ...


class TreeBinded(MyFrame):
    def __init__(self, *args, **kwargs):
        super(TreeBinded, self).__init__(*args, **kwargs)

        self.driver_frame = kwargs.get("driver_frame")
        self.driver_attr = kwargs.get("driver_attr")
        self.rowconfigure(0, weight=1)
        self.rowconfigure(1, weight=0)
        self.columnconfigure(0, weight=1)
        self.columnconfigure(1, weight=0)

        self.tree = ttk.Treeview(master=self, show="tree")
        # self.tree = ttk.Treeview(master=self, show="headings")

        # columns = ("id", )
        # self.tree.configure(columns=columns)

        # self.tree.heading("id", text=LANG.get("id"))

        self.tree.grid(row=0, column=0, sticky=tk.NSEW)

        scroll_y = ttk.Scrollbar(master=self, orient=tk.VERTICAL, command=self.tree.yview)
        scroll_y.grid(row=0, column=1, sticky=tk.NS)
        self.tree.configure(yscrollcommand=scroll_y.set)

        scroll_x = ttk.Scrollbar(master=self, orient=tk.HORIZONTAL, command=self.tree.xview)
        scroll_x.grid(row=1, column=0, sticky=tk.EW)
        self.tree.configure(xscrollcommand=scroll_x.set)

        self.tree.bind('<<TreeviewSelect>>', self.selecting_item)

        self.project = self.nametowidget(".").project

    def show_binds(self, file_id):
        for item in self.tree.get_children():
            self.tree.delete(item)

        self.project = self.nametowidget(".").load_project()
        binds = self.project["files"][file_id].get("binds")
        if binds:
            binds_vertical = []
            binds_horizontal = []
            for bind in binds:
                b = self.project["binds"][str(bind)].get("vertical")
                if b:
                    binds_vertical.append(b)
                else:
                    binds_horizontal.append(self.project["binds"][str(bind)].get("horizontal"))

            print(binds_vertical)
            print(binds_horizontal)

            iid_num = 0
            for bind in binds_vertical:


                ## coś tam działa
                print(bind)
                self.tree.insert("", tk.END, iid=str(iid_num), text=bind[0])
                _iid_num = iid_num + 1
                self.tree.insert(str(iid_num), tk.END, iid=str(_iid_num), text=bind[1])
                iid_num += 2
                # try:
                #     self.tree.insert("",tk.END, iid=, text=bind[0])
                # except tk.TclError:
                #     pass

                # try:
                #     self.tree.insert(bind[0],tk.END, text=bind[1])
                # except tk.TclError:
                #     pass
        else:
            messagebox.showinfo(master=self, title=LANG.get("bind_manager"), message=LANG.get("no_binds"))

    def selecting_item(self, event):
        try:
            for child in self.driver_frame.winfo_children():
                child.destroy()
        except tk.TclError:
            return
        except KeyError:
            return

        item_index = self.tree.focus()
        item_obj = self.tree.item(item_index)
        item_vals = item_obj.get("text")

        try:
            file_id = item_vals
            path_to_screen = self.project["files"][file_id]["path"]
        except KeyError:
            return

        try:
            self.driver_attr = ZoomPan(master=self.driver_frame, path=path_to_screen)
            gc.disable()
            gc.collect()
            gc.enable()
        except tk.TclError:
            return
        except KeyError:
            return


class TreeChoiceOrToBind(MyFrame):
    def __init__(self, *args, **kwargs):
        super(TreeChoiceOrToBind, self).__init__(*args, **kwargs)

        self.driver_frame = kwargs.get("driver_frame")
        self.driver_attr = kwargs.get("driver_attr")

        self.rowconfigure(0, weight=1)
        self.rowconfigure(1, weight=0)
        self.columnconfigure(0, weight=1)
        self.columnconfigure(1, weight=0)

        self.tree = ttk.Treeview(master=self, show="headings")

        columns = ("id", "source", "labels")
        self.tree.configure(columns=columns)

        self.tree.heading("id", text=LANG.get("id"))
        self.tree.heading("source", text=LANG.get("source"))
        self.tree.heading("labels", text=LANG.get("labels"))

        self.tree.grid(row=0, column=0, sticky=tk.NSEW)

        scroll_y = ttk.Scrollbar(master=self, orient=tk.VERTICAL, command=self.tree.yview)
        scroll_y.grid(row=0, column=1, sticky=tk.NS)
        self.tree.configure(yscrollcommand=scroll_y.set)

        scroll_x = ttk.Scrollbar(master=self, orient=tk.HORIZONTAL, command=self.tree.xview)
        scroll_x.grid(row=1, column=0, sticky=tk.EW)
        self.tree.configure(xscrollcommand=scroll_x.set)

        self.tree.bind('<<TreeviewSelect>>', self.selecting_item)

        self.project = self.nametowidget(".").project

    def populate(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        self.project = self.nametowidget(".").load_project()

        for file_id, i in self.project["files"].items():
            self.tree.insert("", tk.END, values=(file_id, i.get("source"), i.get("labels")))

    def selecting_item(self, event):
        try:
            for child in self.driver_frame.winfo_children():
                child.destroy()
        except tk.TclError:
            return
        except KeyError:
            return

        item_index = self.tree.focus()
        item_obj = self.tree.item(item_index)
        item_vals = item_obj.get("values")

        try:
            file_id = item_vals[0]
            path_to_screen = self.project["files"][file_id]["path"]
        except IndexError:
            return

        try:
            self.driver_attr = ZoomPan(master=self.driver_frame, path=path_to_screen)
            self.driver_var[0] = file_id
            print(self.driver_var)

            gc.disable()
            gc.collect()
            gc.enable()
        except tk.TclError:
            return
        except KeyError:
            return


class FilterManager(tk.Toplevel):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.title(LANG.get("filter_labels"))
        self.grab_set()

        self.filter_frame = MyFrame(master=self)
        self.filter_frame.grid()

        self.filter_frame.columnconfigure(0, weight=1)
        self.filter_frame.columnconfigure(1, weight=1)
        self.filter_frame.columnconfigure(2, weight=1)

        self.filter_frame.rowconfigure(0, weight=1)
        self.filter_frame.rowconfigure(1, weight=1)
        self.filter_frame.rowconfigure(2, weight=1)

        ttk.Label(master=self.filter_frame, text=LANG.get("sel_lbls_incl")).grid(row=0, column=0)
        ttk.Label(master=self.filter_frame, text=LANG.get("sel_lbls_excl")).grid(row=0, column=1)

        self.menu_frm = MyFrame(master=self.filter_frame)
        self.menu_frm.grid(row=1, column=2)

        self.project = self.nametowidget(".").load_project()

        labels = self.project["labels"]

        self.lst_box_include = tk.Listbox(master=self.filter_frame, selectmode=tk.MULTIPLE, width=60,
                                          exportselection=False)
        self.lst_box_include.grid(row=1, column=0)
        for l in labels:
            self.lst_box_include.insert(tk.END, l)

        self.lst_box_exclude = tk.Listbox(master=self.filter_frame, selectmode=tk.MULTIPLE, width=60,
                                          exportselection=False)
        self.lst_box_exclude.grid(row=1, column=1)
        for l in labels:
            self.lst_box_exclude.insert(tk.END, l)

        self.operator_labels_include = tk.BooleanVar()
        self.operator_labels_exclude = tk.BooleanVar()

        self.option_frame_include = MyFrame(master=self.filter_frame)
        self.option_frame_include.grid(row=2, column=0)

        ttk.Radiobutton(master=self.option_frame_include, variable=self.operator_labels_include,
                        text=LANG.get("or_include"), value=False).grid()
        ttk.Radiobutton(master=self.option_frame_include, variable=self.operator_labels_include,
                        text=LANG.get("and_include"), value=True).grid()

        self.option_frame_exclude = MyFrame(master=self.filter_frame)
        self.option_frame_exclude.grid(row=2, column=1)

        ttk.Radiobutton(master=self.option_frame_exclude, variable=self.operator_labels_exclude,
                        text=LANG.get("or_exclude"), value=False).grid()
        ttk.Radiobutton(master=self.option_frame_exclude, variable=self.operator_labels_exclude,
                        text=LANG.get("and_exclude"), value=True).grid()

        ttk.Button(master=self.menu_frm, text=LANG.get("filter"), command=self.filter).grid()
        ttk.Button(master=self.menu_frm, text=LANG.get("show_all"), command=self.show_all).grid()
        ttk.Button(master=self.menu_frm, text=LANG.get("go_back"), command=self.destroy).grid()

    def show_all(self):
        self.nametowidget(".!myframe.!baseprojectview").reset_tree()

    def filter(self):
        positive = self.lst_box_include.curselection()
        negative = self.lst_box_exclude.curselection()

        positive = [self.lst_box_include.get(p) for p in positive]
        negative = [self.lst_box_exclude.get(n) for n in negative]

        include = self.operator_labels_include.get()
        exclude = self.operator_labels_exclude.get()

        tree_pan = self.nametowidget(".!myframe.!baseprojectview.!myframe2.!treepan")

        for item in tree_pan.tree.get_children():
            tree_pan.tree.delete(item)

        file_id_filtered_positive = []
        file_id_filtered_negative = []

        if include:
            for file_id, i in self.project["files"].items():
                lbls = i["labels"]
                p = [l for l in lbls if l in positive]
                if len(p) == len(positive):
                    file_id_filtered_positive.append(file_id)

        if exclude:
            for file_id, i in self.project["files"].items():
                lbls = i["labels"]
                n = [l for l in lbls if l in negative]
                if len(n) == len(negative):
                    file_id_filtered_negative.append(file_id)

        if not include:
            for file_id, i in self.project["files"].items():
                lbls = i["labels"]
                p = [l for l in lbls if l in positive]
                if p:
                    file_id_filtered_positive.append(file_id)

        if not exclude:
            for file_id, i in self.project["files"].items():
                lbls = i["labels"]
                n = [l for l in lbls if l in negative]
                if n:
                    file_id_filtered_negative.append(file_id)

        file_id_filtered = [f_id for f_id in file_id_filtered_positive if f_id not in file_id_filtered_negative]

        for file_id in file_id_filtered:
            i = self.project["files"][file_id]
            tree_pan.tree.insert("", tk.END, values=(file_id, i.get("source"), i.get("path"), i.get("labels"),
                                                                 dt.datetime.fromtimestamp(i.get("c_time")).strftime(
                                                                     "%Y-%m-%d %H:%M:%S")))

        gc.disable()
        gc.collect()
        gc.enable()


class LabelManager(tk.Toplevel):
    def __init__(self, *args, **kwargs):
        super(LabelManager, self).__init__(*args, **kwargs)

        self.title(LANG.get("labels"))
        self.grab_set()

        self.frame_labels = ttk.Frame(master=self)
        self.frame_labels.grid()
        self.frame_labels.columnconfigure(0)
        self.frame_labels.columnconfigure(1)

        self.frame_labels.rowconfigure(0)

        self.lst_box = tk.Listbox(master=self.frame_labels, width=60)
        self.lst_box.grid(row=0, column=0)

        self.project = self.nametowidget(".").load_project()

        self.labels = self.project["labels"]
        for l in self.labels:
            self.lst_box.insert(tk.END, l)

        self.right_pan = ttk.Frame(master=self.frame_labels)
        self.right_pan.grid(row=0, column=1)

        ttk.Label(master=self.right_pan, text=LANG.get("input_new_label")).grid()

        self.ent_add = ttk.Entry(master=self.right_pan, width=60)
        self.ent_add.grid()

        ttk.Button(master=self.right_pan, text=LANG.get("add_label"), command=self.add_label).grid()
        ttk.Button(master=self.right_pan, text=LANG.get("del_label"), command=self.del_label).grid()
        ttk.Button(master=self.right_pan, text=LANG.get("edit_label"), command=self.edit_label).grid()
        ttk.Button(master=self.right_pan, text=LANG.get("go_back"), command=self.destroy).grid()

    def add_label(self):
        temp_lbl = self.ent_add.get()
        if temp_lbl and temp_lbl not in self.labels:
            self.project["labels"].append(temp_lbl)
            self.lst_box.insert(tk.END, temp_lbl)

            self.nametowidget('.').save_project()
            self.project = self.nametowidget(".").load_project()
        else:
            messagebox.showerror(master=self, title=LANG.get("labels"), message=LANG.get("no_labels"))

    def del_label(self):
        temp_lbl = self.lst_box.curselection()
        if temp_lbl:
            num_of_lbl = temp_lbl[0]
            temp_lbl = self.lst_box.get(num_of_lbl)
            counter = 0
            for file, val in self.project["files"].items():
                if temp_lbl in val.get("labels"):
                    counter += 1

            res = messagebox.askyesno(master=self, title=LANG.get("labels"),
                                      message=f'{LANG.get("ask_del_conf_1")}{counter}{LANG.get("ask_del_conf_2")}{temp_lbl}')

            if res:
                self.lst_box.delete(num_of_lbl)

                for file, val in self.project["files"].items():
                    if temp_lbl in val.get("labels"):
                        val.get("labels").remove(temp_lbl)

                self.project["labels"].remove(temp_lbl)

                self.nametowidget('.').save_project()
                self.project = self.nametowidget(".").load_project()
                self.nametowidget(".!myframe.!baseprojectview").reset_tree()
            else:
                return

        else:
            messagebox.showerror(master=self, title=LANG.get("labels"), message=LANG.get("no_label_selected"))

    def edit_label(self):

        self.temp_lbl = self.lst_box.curselection()
        if self.temp_lbl:
            self.num_of_lbl = self.temp_lbl[0]
            self.temp_lbl = self.lst_box.get(self.temp_lbl[0])

            self.edit_frame = ttk.Frame(master=self.right_pan)
            self.edit_frame.grid()

            ttk.Label(master=self.edit_frame, text=f'{LANG.get("rename_label")}{self.temp_lbl}').grid()

            self.ent_edit = ttk.Entry(master=self.edit_frame, width=60)
            self.ent_edit.grid()

            ttk.Button(master=self.edit_frame, text=LANG.get("submit"), command=self._edit).grid()
            ttk.Button(master=self.edit_frame, text=LANG.get("go_back"), command=self.edit_frame.destroy).grid()
        else:
            messagebox.showerror(master=self, title=LANG.get("labels"), message=LANG.get("no_label_selected"))

    def _edit(self):
        new_lbl = self.ent_edit.get()

        if not new_lbl:
            messagebox.showerror(master=self, title=LANG.get("labels"), message=LANG.get("no_new_label_name"))
            self.edit_frame.destroy()
            return

        counter = 0
        for file, val in self.project["files"].items():
            if self.temp_lbl in val.get("labels"):
                counter += 1

        res = messagebox.askyesno(master=self, title=LANG.get("labels"),
                                  message=f'{LANG.get("ask_edit_conf_1")}{counter}{LANG.get("ask_edit_conf_2")}{self.temp_lbl}')

        if res:
            self.lst_box.delete(self.num_of_lbl)
            self.lst_box.insert(self.num_of_lbl, new_lbl)

            for file, val in self.project["files"].items():
                if self.temp_lbl in val.get("labels"):
                    index = val.get("labels").index(self.temp_lbl)
                    val.get("labels")[index] = new_lbl

            index = self.project["labels"].index(self.temp_lbl)
            self.project["labels"][index] = new_lbl

            self.nametowidget('.').save_project()
            self.project = self.nametowidget(".").load_project()

            self.edit_frame.destroy()
            self.master.reset_tree()
        else:
            return


class SourceManager(tk.Toplevel):
    def __init__(self, *args, **kwargs):
        super(SourceManager, self).__init__(*args, **kwargs)

        self.title(f'{LANG.get("manage_links")}')
        self.columnconfigure(0, weight=0, minsize=300)
        self.columnconfigure(1, weight=1, minsize=500)
        self.columnconfigure(2, weight=1, minsize=500)
        self.rowconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)
        # self.rowconfigure(2, weight=1)
        self.grab_set()

        # style = ttk.Style()
        # style.configure("BW.TLabel", background="black")

        self.menu = MyFrame(master=self)
        # self.menu = MyFrame(master=self, style="BW.TLabel")

        self.menu.grid(column=0, row=0)

        self.menu.columnconfigure(0, weight=1)
        self.menu.rowconfigure((0, 1, 2), weight=0)
        ttk.Button(master=self.menu, text=LANG.get("add_links"),
                   command=self.add_sources).grid(sticky=tk.EW)

        ttk.Button(master=self.menu, text=LANG.get("assign_sources"),
                   command=self.assign_sources).grid(sticky=tk.EW)

        ttk.Button(master=self.menu, text=LANG.get("go_back"), command=self.destroy).grid(sticky=tk.EW)

        self.temp_upper_pan = MyFrame(master=self)
        self.temp_upper_pan.grid(column=1, row=0, columnspan=2, )
        self.temp_upper_pan.columnconfigure(0, weight=1)
        self.temp_upper_pan.rowconfigure(0, weight=1)

        self.temp_left_pan = MyFrame(master=self)
        self.temp_left_pan.grid(column=1, row=1)
        self.temp_left_pan.columnconfigure(0, weight=1)
        self.temp_left_pan.columnconfigure(1, weight=0)
        self.temp_left_pan.rowconfigure(0, weight=1)
        self.temp_left_pan.rowconfigure(1, weight=0)

        self.temp_right_pan = MyFrame(master=self)
        self.temp_right_pan.grid(column=2, row=1)
        self.temp_right_pan.columnconfigure(0, weight=1)
        self.temp_right_pan.columnconfigure(1, weight=0)
        self.temp_right_pan.rowconfigure(0, weight=1)
        self.temp_right_pan.rowconfigure(1, weight=0)

        self.temp_bottom_pan = MyFrame(master=self)
        self.temp_bottom_pan.grid(row=1, column=0)
        self.temp_bottom_pan.rowconfigure((0, 1, 2, 3, 4), weight=0)
        self.temp_bottom_pan.columnconfigure(0, weight=1)

        self.pat_wspace = re.compile("^\s+$")

        self.tbox = None

    def add_sources(self):
        for child in self.temp_upper_pan.winfo_children():
            child.destroy()

        for child in self.temp_left_pan.winfo_children():
            child.destroy()

        for child in self.temp_right_pan.winfo_children():
            child.destroy()

        for child in self.temp_bottom_pan.winfo_children():
            child.destroy()

        self.tbox_pan = MyFrame(master=self.temp_upper_pan)
        self.tbox_pan.grid()
        self.tbox_pan.columnconfigure(0, weight=1)
        self.tbox_pan.rowconfigure(0, weight=1)
        self.tbox_pan.rowconfigure(1, weight=0)

        self.tbox = scrolledtext.ScrolledText(master=self.tbox_pan)
        self.tbox.grid(row=0, column=0, sticky=tk.NSEW)

        self.tbox.insert(tk.END, LANG.get("input_links"))

        ttk.Button(master=self.tbox_pan, text=LANG.get("submit"), command=self.accept_sources).grid(row=1, column=0,
                                                                                                    padx=5, pady=2)

    def accept_sources(self):
        links = [x for x in self.tbox.get("1.0", tk.END).split("\n") if bool(x) and not self.pat_wspace.match(x)]

        if not links:
            messagebox.showerror(master=self, title=LANG.get("manage_links"),  message=LANG.get("no_links_or_wrong"))
            return

        self.project = self.nametowidget(".").load_project()

        if "links" in self.project:
            self.project["links"] += links
        else:
            self.project["links"] = links

        self.nametowidget(".").project = self.project
        self.nametowidget(".").save_project()

        messagebox.showinfo(master=self, title=LANG.get("manage_links"), message=f"{LANG.get('source_added')}{len(links)}")

        self.add_sources()

    def assign_sources(self):
        for child in self.temp_upper_pan.winfo_children():
            child.destroy()

        for child in self.temp_left_pan.winfo_children():
            child.destroy()

        for child in self.temp_right_pan.winfo_children():
            child.destroy()

        for child in self.temp_bottom_pan.winfo_children():
            child.destroy()

        self.project = self.nametowidget(".").load_project()

        self.all_assigned_links = list(set(v.get("source") for k, v in self.project["files"].items()))

        self.copy_project_files = copy.deepcopy(self.project["files"])

        self.selected_source = None
        self.selected_base = None
        self.changes_applied = None
        self.selected_sources_to_delete = []

        self.tree_sources = ttk.Treeview(master=self.temp_left_pan, show="headings")
        columns_scr = ("source",)
        self.tree_sources.configure(columns=columns_scr)
        self.tree_sources.column("source", width=500)
        self.tree_sources.heading("source", text=LANG.get("source"))
        self.tree_sources.grid(row=0, column=0, sticky=tk.NSEW)

        scroll_y_scr = ttk.Scrollbar(master=self.temp_left_pan, orient=tk.VERTICAL, command=self.tree_sources.yview)
        scroll_y_scr.grid(row=0, column=1, sticky=tk.NS)
        self.tree_sources.configure(yscrollcommand=scroll_y_scr.set)

        scroll_x_src = ttk.Scrollbar(master=self.temp_left_pan, orient=tk.HORIZONTAL, command=self.tree_sources.xview)
        scroll_x_src.grid(row=1, column=0, sticky=tk.EW)
        self.tree_sources.configure(xscrollcommand=scroll_x_src.set)

        self.tree_sources.bind('<<TreeviewSelect>>', self.selecting_sources)

        for link in self.project["links"]:
            self.tree_sources.insert("", tk.END, value=(link,))

        self.tree_base = ttk.Treeview(master=self.temp_right_pan, show="headings")
        columns_base = ("id", "source", "path", "labels", "c_time")

        self.tree_base.configure(columns=columns_base)

        self.tree_base.heading("id", text=LANG.get("id"))
        self.tree_base.heading("source", text=LANG.get("source"))
        self.tree_base.heading("path", text=LANG.get("path"))
        self.tree_base.heading("labels", text=LANG.get("labels"))
        self.tree_base.heading("c_time", text=LANG.get("c_time"))

        self.tree_base.grid(row=0, column=0, sticky=tk.NSEW)

        scroll_y = ttk.Scrollbar(master=self.temp_right_pan, orient=tk.VERTICAL, command=self.tree_base.yview)
        scroll_y.grid(row=0, column=1, sticky=tk.NS)
        self.tree_base.configure(yscrollcommand=scroll_y.set)

        scroll_x = ttk.Scrollbar(master=self.temp_right_pan, orient=tk.HORIZONTAL, command=self.tree_base.xview)
        scroll_x.grid(row=1, column=0, sticky=tk.EW)
        self.tree_base.configure(xscrollcommand=scroll_x.set)

        self.tree_base.bind('<<TreeviewSelect>>', self.selecting_base)

        for file_id, i in self.copy_project_files.items():
            if i.get("source") == "":
                self.tree_base.insert("", tk.END, values=(file_id, i.get("source"), i.get("path"), i.get("labels"),
                                                          dt.datetime.fromtimestamp(i.get("c_time")).strftime(
                                                              "%Y-%m-%d %H:%M:%S")))

        btn_show_all_links = ttk.Button(master=self.temp_bottom_pan, text=LANG.get("show_all_links"),
                                        command=self.show_all_links)
        btn_show_all_links.grid(row=0, column=0, sticky=tk.NSEW)

        btn_show_all_links = ttk.Button(master=self.temp_bottom_pan, text=LANG.get("show_unassigned_links"),
                                        command=self.show_unassigned_links)
        btn_show_all_links.grid(row=1, column=0, sticky=tk.NSEW)

        btn_assign = ttk.Button(master=self.temp_bottom_pan, text=LANG.get("assign"), command=self.assign)
        btn_assign.grid(row=2, column=0, sticky=tk.NSEW)

        btn_save = ttk.Button(master=self.temp_bottom_pan, text=LANG.get("save_changes"),
                              command=self.save_changes)
        btn_save.grid(row=3, column=0, sticky=tk.NSEW)

        btn_delete_source = ttk.Button(master=self.temp_bottom_pan, text=LANG.get("del_source"),
                                       command=self.delete_source)
        btn_delete_source.grid(row=4, column=0, sticky=tk.NSEW)

    def show_all_links(self):

        # all_assigned_links = list(set(v.get("source") for k, v in project["files"].items()))
        unassigned_links = [self.tree_sources.item(i).get("values")[0] for i in self.tree_sources.get_children()]

        for link in self.all_assigned_links:
            if link not in unassigned_links:
                self.tree_sources.insert("", tk.END, value=(link,))

    def show_unassigned_links(self):
        for i in self.tree_sources.get_children():
            if self.tree_sources.item(i).get("values")[0] not in self.project["links"]:
                self.tree_sources.delete(i)

    def selecting_sources(self, event):

        curr_item = self.tree_sources.focus()
        item = self.tree_sources.item(curr_item)
        self.selected_source = (curr_item, item.get("values")[0])

        print(self.selected_source)

    def selecting_base(self, event):

        for child in self.temp_upper_pan.winfo_children():
            child.destroy()

        curr_item = self.tree_base.focus()
        item = self.tree_base.item(curr_item)

        ZoomAdvanced(master=self.temp_upper_pan, path=pathlib.Path(item.get("values")[2])).grid(sticky=tk.NSEW)

        self.selected_base = (curr_item, item.get("values")[0])

        print(self.selected_base)

    def assign(self):
        if self.selected_source and self.selected_base:
            res = messagebox.askyesno(master=self, title=LANG.get("assign_sources"), message=LANG.get("ask_assign"))

            if res:
                key_temp = self.selected_base[1]
                self.copy_project_files[key_temp]["source"] = self.selected_source[1]

                self.tree_base.item(self.selected_base[0], values=(key_temp,
                                                                   self.copy_project_files[key_temp].get("source"),
                                                                   self.copy_project_files[key_temp].get("path"),
                                                                   self.copy_project_files[key_temp].get("labels"),
                                                                   dt.datetime.fromtimestamp(
                                                                       self.copy_project_files[key_temp].get(
                                                                           "c_time")).strftime(
                                                                       "%Y-%m-%d %H:%M:%S")))

                self.changes_applied = True
                self.selected_sources_to_delete.append(self.selected_source[1])
            else:
                return

        else:
            messagebox.showinfo(master=self, title=LANG.get("assign_sources"), message=LANG.get("mark_sources_and_img"))

    def save_changes(self):
        if not self.changes_applied:
            messagebox.showinfo(master=self, title=LANG.get("assign_sources"), message=LANG.get("no_changes_applied"))
            return

        res = messagebox.askyesno(master=self, title=LANG.get("assign_sources"), message=LANG.get("ask_save_changes"))
        if res:
            self.project["files"] = self.copy_project_files

            for scr_to_del in self.selected_sources_to_delete:
                try:
                    self.project["links"].remove(scr_to_del)
                except:
                    pass

            self.nametowidget(".").project = self.project
            self.nametowidget(".").save_project()

            self.nametowidget(".!myframe.!baseprojectview").reset_tree()

            self.changes_applied = False
            self.selected_sources_to_delete = []

            self.assign_sources()

    def delete_source(self):
        # todo - czy jeśli wywalam a source jest przypisane to też wywalać? czy wywalać tylko nieprzypisane
        if not self.selected_source:
            messagebox.showinfo(master=self, title=LANG.get("del_source"), message=LANG.get("no_source_sel"))
            return

        self.selected_sources_to_delete.append(self.selected_source[1])
        self.changes_applied = True
        messagebox.showinfo(master=self, title=LANG.get("del_source"), message=LANG.get("save_to_confirm_del"))


class FieldsManager(tk.Toplevel):
    def __init__(self, driver, file_id, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.title(LANG.get("fields_manager"))

        self.grab_set()

        self.m_frame = ttk.Frame(master=self)
        self.m_frame.grid()

        self.m_frame.columnconfigure(0)
        self.m_frame.columnconfigure(1)
        self.m_frame.columnconfigure(2)

        self.m_frame.rowconfigure(0)

        self.left_pan = ttk.Frame(master=self.m_frame)
        self.left_pan.grid(column=0, row=0)

        self.centr_pan = ttk.Frame(master=self.m_frame)
        self.centr_pan.grid(column=1, row=0)

        self.right_pan = ttk.Frame(master=self.m_frame)
        self.right_pan.grid(column=2, row=0)

        ttk.Label(master=self.centr_pan, text=f'{LANG.get("curr_fields")}{file_id}').grid()

        self.lst_box_fields = tk.Listbox(master=self.centr_pan, width=60)
        self.lst_box_fields.grid()

        self.project = self.nametowidget(".").load_project()
        self.extra_fields = self.project["files"][file_id]["extra_fields"]
        self.file_id = file_id
        self.driver = driver

        for f, val in self.extra_fields.items():
            self.lst_box_fields.insert(tk.END, f"{f} ___ {val}")

        ttk.Button(master=self.left_pan, text=LANG.get("add_field"), command=self.add_field).grid()
        ttk.Button(master=self.left_pan, text=LANG.get("edit_field"), command=self.edit_field).grid()
        ttk.Button(master=self.left_pan, text=LANG.get("del_field"), command=self.del_field).grid()
        ttk.Button(master=self.left_pan, text=LANG.get("go_back"), command=self.destroy).grid()

    def add_field(self):
        for c in self.right_pan.winfo_children():
            c.destroy()

        ttk.Label(master=self.right_pan, text=LANG.get("add_field")).grid()

        ttk.Label(master=self.right_pan, text=LANG.get("field_name")).grid()
        self.field_name = ttk.Entry(master=self.right_pan)
        self.field_name.grid()

        ttk.Label(master=self.right_pan, text=LANG.get("field_val")).grid()
        self.field_val = ttk.Entry(master=self.right_pan)
        self.field_val.grid()

        ttk.Button(master=self.right_pan, text=LANG.get("submit"),
                   command=self.accept_add).grid()

    def accept_add(self):
        f_name = self.field_name.get()
        if f_name:
            extra_fields = self.project["files"][self.file_id]["extra_fields"]

            if f_name in extra_fields:
                messagebox.showerror(master=self, title=LANG.get("field_name_exists"), message=LANG.get("field_name_exists"))
                return

            extra_fields[f_name] = self.field_val.get()

            self.nametowidget(".").project = self.project
            self.nametowidget(".").save_project()

            self.lst_box_fields.insert(tk.END, f"{f_name} ___ {self.field_val.get()}")

            for c in self.driver.frame_extra.winfo_children():
                c.destroy()

            for index, (field_name, field_val) in enumerate(self.extra_fields.items()):
                ExtraField(master=self.driver.frame_extra, field_name=field_name, field_val=field_val, file_id=self.file_id
                           ).grid(row=index // 3, column=index % 3)

            for c in self.right_pan.winfo_children():
                c.destroy()

    def edit_field(self):
        for c in self.right_pan.winfo_children():
            c.destroy()

        self.f_index = self.lst_box_fields.curselection()
        if self.f_index:
            for c in self.right_pan.winfo_children():
                c.destroy()

            self.f_index = self.f_index[0]
            self.f_edit_name, self.f_edit_val = self.lst_box_fields.get(self.f_index).split(" ___ ")

            ttk.Label(master=self.right_pan, text=LANG.get("edit_field")).grid()

            ttk.Label(master=self.right_pan, text=LANG.get("rename_field")).grid()
            self.new_fname_ent = ttk.Entry(master=self.right_pan)
            self.new_fname_ent.insert(tk.END, self.f_edit_name)
            self.new_fname_ent.grid()

            ttk.Label(master=self.right_pan, text=LANG.get("change_val_field")).grid()
            self.new_val_ent = ttk.Entry(master=self.right_pan)
            self.new_val_ent.insert(tk.END, self.f_edit_val)
            self.new_val_ent.grid()

            ttk.Button(master=self.right_pan, text=LANG.get("submit"),
                       command=self.accept_edit).grid()

    def accept_edit(self):

        if self.new_fname_ent.get() != self.f_edit_name and self.new_val_ent.get() != self.f_edit_val:
            self.project["files"][self.file_id]["extra_fields"][self.new_fname_ent.get()] = self.new_val_ent.get()

            del self.project["files"][self.file_id]["extra_fields"][self.f_edit_name]

            self.lst_box_fields.delete(self.f_index)
            self.lst_box_fields.insert(self.f_index, f"{self.new_fname_ent.get()} ___ {self.new_val_ent.get()}")

        elif self.new_fname_ent.get() != self.f_edit_name:
            self.project["files"][self.file_id]["extra_fields"][self.new_fname_ent.get()] = self.f_edit_val

            del self.project["files"][self.file_id]["extra_fields"][self.f_edit_name]

            self.lst_box_fields.delete(self.f_index)
            self.lst_box_fields.insert(self.f_index, f"{self.new_fname_ent.get()} ___ {self.f_edit_val}")

        else:
            self.project["files"][self.file_id]["extra_fields"][self.f_edit_name] = self.new_val_ent.get()
            self.lst_box_fields.delete(self.f_index)
            self.lst_box_fields.insert(self.f_index, f"{self.new_fname_ent.get()} ___ {self.new_val_ent.get()}")

        self.nametowidget(".").project = self.project
        self.nametowidget(".").save_project()

        for c in self.driver.frame_extra.winfo_children():
            c.destroy()

        for index, (field_name, field_val) in enumerate(self.extra_fields.items()):
            ExtraField(master=self.driver.frame_extra, field_name=field_name, field_val=field_val, file_id=self.file_id
                       ).grid(row=index // 3, column=index % 3)

        for c in self.right_pan.winfo_children():
            c.destroy()

    def del_field(self):
        for c in self.right_pan.winfo_children():
            c.destroy()

        self.f_index = self.lst_box_fields.curselection()
        if self.f_index:
            res = messagebox.askyesno(master=self, title=LANG.get("fields_manager"), message=LANG.get("del_field"))
            if res:
                self.f_index = self.f_index[0]
                self.f_del_name, _ = self.lst_box_fields.get(self.f_index).split(" ___ ")

                del self.project["files"][self.file_id]["extra_fields"][self.f_del_name]

                self.nametowidget(".").project = self.project
                self.nametowidget(".").save_project()

                self.lst_box_fields.delete(self.f_index)

                for c in self.driver.frame_extra.winfo_children():
                    c.destroy()

                for index, (field_name, field_val) in enumerate(self.extra_fields.items()):
                    ExtraField(master=self.driver.frame_extra, field_name=field_name, field_val=field_val,
                               file_id=self.file_id
                               ).grid(row=index // 3, column=index % 3)


class DetailPan(MyFrame):
    def __init__(self, driver, file_id, item_index, *args, **kwargs):

        super(DetailPan, self).__init__(*args, **kwargs)
        print(self.nametowidget(self))

        self.driver = driver
        self.item_index = item_index
        self.file_id = file_id

        self.columnconfigure(0)
        self.columnconfigure(1)
        self.columnconfigure(2)

        for row in range(8):
            self.rowconfigure(row)

        self.grid()

        self.project = self.nametowidget(".").load_project()

        self.source = self.project["files"][self.file_id]["source"]
        self.path = self.project["files"][self.file_id]["path"]
        self.labels = self.project["files"][self.file_id]["labels"]
        self.comment = self.project["files"][self.file_id]["comment"]
        self.extra_fields = self.project["files"][self.file_id]["extra_fields"]
        self.c_time = self.project["files"][self.file_id]["c_time"]
        c_time = dt.datetime.fromtimestamp(self.c_time).strftime("%Y-%m-%d %H:%M:%S")
        if self.project["files"][self.file_id].get("binds"):
            self.binds = self.project["files"][self.file_id].get("binds")
        else:
            self.binds = []
        # c_time = f"{self.c_time:%Y-%m-%d %H-%M-%S}"

        ttk.Label(master=self, text=LANG.get("id")).grid(row=0, column=0)

        self.id_pan = MyFrame(master=self)
        self.id_pan.grid(row=0, column=1)

        self.id, self.f_name = self.file_id.split(" - ", 1)
        ttk.Label(master=self.id_pan, text=f"{self.id} - ").pack(side=tk.LEFT)

        self.ent_name = ttk.Entry(master=self.id_pan, width=100 - len(f"{self.id} - "))
        self.ent_name.insert(tk.END, self.f_name)
        self.ent_name.pack(side=tk.LEFT)

        ttk.Button(master=self, text=LANG.get("f_rename"), command=self.rename_file).grid(row=0, column=2)

        ttk.Label(master=self, text=LANG.get("source")).grid(row=1, column=0)
        self.ent_src = ttk.Entry(master=self, width=100)
        self.ent_src.insert(tk.END, self.source)
        self.ent_src.grid(row=1, column=1)

        ttk.Button(master=self, text=LANG.get("s_rename"), command=self.rename_source).grid(row=1, column=2)

        ttk.Label(master=self, text=LANG.get("path")).grid(row=2, column=0)
        path_broken = self.path
        if len(self.path) > 100:
            path_broken = path_broken[:99] + "\n    " + path_broken[99:]
        self.lbl_path = ttk.Label(master=self, text=path_broken, width=100)
        self.lbl_path.grid(row=2, column=1)

        ttk.Label(master=self, text=LANG.get("labels")).grid(row=3, column=0)
        self.frame_lbls = MyFrame(master=self)
        self.frame_lbls.grid(row=3, column=1)

        ttk.Button(master=self, text=LANG.get("manage_labels"), command=self.manage_labels).grid(row=3, column=2)

        for lbl in self.labels:
            ttk.Label(master=self.frame_lbls, text=lbl).pack(side=tk.LEFT, padx=5, pady=2)

        ttk.Label(master=self, text=LANG.get("comment")).grid(row=4, column=0)

        self.tbox_com = scrolledtext.ScrolledText(master=self, height=4)
        self.tbox_com.insert("1.0", self.comment)
        self.tbox_com.grid(row=4, column=1)

        ttk.Button(master=self, text=LANG.get("alter_comment"), command=self.alter_comment).grid(row=4, column=2)

        ttk.Label(master=self, text=LANG.get("extra_fields")).grid(row=5, column=0)
        self.frame_extra = MyFrame(master=self)
        self.frame_extra.grid(row=5, column=1)

        self.frame_extra.columnconfigure(0)
        self.frame_extra.columnconfigure(1)
        self.frame_extra.columnconfigure(2)

        num_fields = len(self.extra_fields)

        for r in range(num_fields):
            if r % 3 == 0:
                self.frame_extra.rowconfigure(r // 3)

        for index, (field_name, field_val) in enumerate(self.extra_fields.items()):
            ExtraField(master=self.frame_extra, field_name=field_name, field_val=field_val, file_id=self.file_id
                       ).grid(row=index // 3, column=index % 3)

        ttk.Button(master=self, text=LANG.get("manage_fields"),
                   command=self.manage_fields).grid(row=5, column=2)

        ttk.Label(master=self, text=LANG.get("c_time")).grid(row=6, column=0)
        lbl_ctime = ttk.Label(master=self, text=c_time, width=100)
        lbl_ctime.grid(row=6, column=1)

        ttk.Button(master=self, text=LANG.get("del_file"), command=self.del_file).grid(row=7, column=1)  # todo command

        self.fields_manager = None

    # todo: add "open in new window option"
    def del_file(self):
        binds_affected = [bind for bind in self.project["binds"] if self.file_id in bind]
        res = messagebox.askyesno(master=self, title=LANG.get("del_file"), message=f"{LANG.get('del_file_confirm')}\n{self.file_id}\n{LANG.get('binds_affected')}{len(binds_affected)}")
        if res:
            f_path = pathlib.Path(self.path)
            f_path.unlink()

            self.project["binds"] = [bind for bind in self.project["binds"] if self.file_id not in bind]

            del self.project["files"][self.file_id]

            self.nametowidget(".").project = self.project
            self.nametowidget(".").save_project()

            tree_pan = self.nametowidget(".!myframe.!baseprojectview.!myframe2.!treepan")
            tree_pan.project = self.project

            tree_pan.tree.delete(self.item_index)

            for c in self.nametowidget(".!myframe.!baseprojectview.!myframe3").winfo_children():
                print(self.nametowidget(c))
                c.destroy()

            self.destroy()

    def manage_fields(self):
        self.fields_manager = FieldsManager(driver=self, file_id=self.file_id)

    def rename_file(self):
        n_name = self.ent_name.get()

        if not n_name.endswith(pathlib.Path(self.path).suffix):
            messagebox.showerror(master=self, title=LANG.get("f_rename"), message=LANG.get("f_name_wrong_suffix"))
            return

        if n_name != self.f_name:
            res = messagebox.askyesno(master=self, title=LANG.get("f_rename"), message=LANG.get("f_rename"))
            if res:
                old_path = pathlib.Path(self.path)
                new_path = pathlib.Path(old_path.parent, n_name)
                old_path.rename(new_path)

                self.path = str(new_path)

                path_broken = self.path
                if len(self.path) > 100:
                    path_broken = path_broken[:99] + "\n    " + path_broken[99:]
                self.lbl_path.configure(text=path_broken)

                self.f_name = n_name

                self.project["files"][f"{self.id} - {self.f_name}"] = {
                    "id": int(self.id),
                    "source": self.source,
                    "path": self.path,
                    "labels": self.labels,
                    "comment": self.comment,
                    "extra_fields": self.extra_fields,
                    "c_time": self.c_time,
                    "binds": self.binds
                }

                del self.project["files"][self.file_id]

                self.project["files"] = dict(sorted(self.project["files"].items(), key=lambda item: item[1]["id"]))

                self.nametowidget(".!myframe.!baseprojectview.!myframe2.!treepan").project = self.project
                self.nametowidget(".").save_project()

                self.file_id = f"{self.id} - {self.f_name}"

                tree_pan = self.nametowidget(".!myframe.!baseprojectview.!myframe2.!treepan")
                tree_pan.project = self.project
                tree_pan.tree.item(
                    self.item_index,
                    values=(
                        (self.file_id,
                         self.source,
                         self.path,
                         self.labels,
                         dt.datetime.fromtimestamp(self.c_time).strftime("%Y-%m-%d %H:%M:%S"))
                    )
                )
        else:
            messagebox.showerror(master=self, title=LANG.get("f_rename"), message=LANG.get("same_f_name"))

    def rename_source(self):
        src = self.ent_src.get()
        if src != self.source:
            res = messagebox.askyesno(master=self, title=LANG.get("s_rename"), message=LANG.get("s_rename"))
            if res:
                self.project["files"][self.file_id]["source"] = src
                self.source = src
                self.nametowidget(".").save_project()

                self.nametowidget(".!myframe.!baseprojectview.!myframe2.!treepan").project = self.project

                tree_pan = self.nametowidget(".!myframe.!baseprojectview.!myframe2.!treepan")
                tree_pan.project = self.project
                tree_pan.tree.item(
                    self.item_index,
                    values=(
                        (self.file_id,
                         self.source,
                         self.path,
                         self.labels,
                         dt.datetime.fromtimestamp(self.c_time).strftime("%Y-%m-%d %H:%M:%S"))
                    )
                )

        else:
            messagebox.showerror(master=self, title=LANG.get("s_rename"), message=LANG.get("same_source"))

    def manage_labels(self):

        self.lbl_manager = tk.Toplevel(master=self)
        self.lbl_manager.title(f"{LANG.get('label_man')}{self.file_id}")

        self.lbl_manager.grab_set()

        self.frame_lbl = MyFrame(master=self.lbl_manager)
        self.frame_lbl.grid()

        self.frame_lbl.columnconfigure(0)
        self.frame_lbl.columnconfigure(1)
        self.frame_lbl.columnconfigure(2)

        self.frame_lbl.rowconfigure(0)
        self.frame_lbl.rowconfigure(1)

        self.btn_menu = MyFrame(master=self.frame_lbl)
        self.btn_menu.grid(row=1, column=2, )

        ttk.Label(master=self.frame_lbl, text=LANG.get("all_lbls")).grid(row=0, column=0)
        ttk.Label(master=self.frame_lbl, text=LANG.get("file_lbls")).grid(row=0, column=1)

        self.lst_box_all = tk.Listbox(master=self.frame_lbl, width=60)
        self.lst_box_all.grid(row=1, column=0, sticky=tk.NSEW)

        all_labels = self.project["labels"]
        for l in all_labels:
            self.lst_box_all.insert(tk.END, l)

        self.lst_box_file = tk.Listbox(master=self.frame_lbl, width=60)
        self.lst_box_file.grid(row=1, column=1, sticky=tk.NSEW)

        self.project = self.nametowidget(".").load_project()

        file_labels = self.project["files"][self.file_id]["labels"]
        for l in file_labels:
            self.lst_box_file.insert(tk.END, l)

        ttk.Button(master=self.btn_menu, text=LANG.get("sel_lbl"), command=self._select_lbl).grid(row=0)
        ttk.Button(master=self.btn_menu, text=LANG.get("remove_lbl"), command=self._remove_lbl).grid(row=1)
        ttk.Button(master=self.btn_menu, text=LANG.get("confirm_sel"),
                   command=self._confirm_sel).grid(row=2)
        ttk.Button(master=self.btn_menu, text=LANG.get("go_back"), command=self.lbl_manager.destroy).grid(row=3)
        ttk.Button(master=self.btn_menu, text=LANG.get("go_to_labels"),
                   command=self._go_to_label_manager).grid(row=4)

    def _go_to_label_manager(self):
        self.lbl_manager.destroy()
        self.nametowidget(".!myframe.!baseprojectview").label_manager()

    def _select_lbl(self):
        sel_lbl = self.lst_box_all.curselection()
        if sel_lbl and not self.lst_box_all.get(sel_lbl[0]) in self.lst_box_file.get(0, tk.END):
            self.lst_box_file.insert(tk.END, self.lst_box_all.get(sel_lbl[0]))

    def _remove_lbl(self):
        rem_lbl = self.lst_box_file.curselection()
        if rem_lbl:
            self.lst_box_file.delete(rem_lbl[0])

    def _confirm_sel(self):
        selected_labels = self.lst_box_file.get(0, tk.END)
        print(selected_labels)
        self.project["files"][self.file_id]["labels"] = list(selected_labels)
        print(self.project["files"][self.file_id]["labels"])

        self.nametowidget(".").save_project()
        self.project = self.nametowidget(".").load_project()

        self.labels = self.project["files"][self.file_id]["labels"]

        for child in self.frame_lbls.winfo_children():
            child.destroy()

        for lbl in self.labels:
            ttk.Label(master=self.frame_lbls, text=lbl).pack(side=tk.LEFT, padx=5, pady=2)

        self.nametowidget(".!myframe.!baseprojectview.!myframe2.!treepan").project = self.project

        selected_labels = " ".join(selected_labels)
        tree_pan = self.nametowidget(".!myframe.!baseprojectview.!myframe2.!treepan")
        tree_pan.project = self.project
        tree_pan.tree.item(
            self.item_index,
            values=(
                (self.file_id,
                 self.source,
                 self.path,
                 selected_labels,
                 dt.datetime.fromtimestamp(self.c_time).strftime("%Y-%m-%d %H:%M:%S"))
            )
        )

    def alter_comment(self):
        comment = self.tbox_com.get("1.0", tk.END)[:-1]

        if comment != self.comment:
            res = messagebox.askyesno(master=self, title=LANG.get("alter_comment"), message=LANG.get("alter_comment"))
            if res:
                self.project["files"][self.file_id]["comment"] = comment
                self.comment = comment
                self.nametowidget(".").save_project()
        else:
            messagebox.showerror(master=self, title=LANG.get("alter_comment"), message=LANG.get("same_comment"))


class ZoomPan(MyFrame):
    def __init__(self, path=None, *args, **kwargs):
        super(ZoomPan, self).__init__(*args, **kwargs)

        if not path:
            return

        self.path = path

        self.grid(sticky=tk.NSEW)

        self.zoom = ZoomAdvanced(master=self, path=pathlib.Path(self.path))
        self.zoom.grid(sticky=tk.NSEW)


class TreePan(MyFrame):
    def __init__(self, *args, **kwargs):
        super(TreePan, self).__init__(*args, **kwargs)

        self.driver = kwargs.get("driver")

        self.rowconfigure(0, weight=1)
        self.rowconfigure(1, weight=0)
        self.columnconfigure(0, weight=1)
        self.columnconfigure(1, weight=0)

        self.tree = ttk.Treeview(master=self, show="headings")

        columns = ("id", "source", "path", "labels", "c_time")
        self.tree.configure(columns=columns)

        self.tree.heading("id", text=LANG.get("id"))
        self.tree.heading("source", text=LANG.get("source"))
        self.tree.heading("path", text=LANG.get("path"))
        self.tree.heading("labels", text=LANG.get("labels"))
        self.tree.heading("c_time", text=LANG.get("c_time"))

        self.tree.grid(row=0, column=0, sticky=tk.NSEW)

        scroll_y = ttk.Scrollbar(master=self, orient=tk.VERTICAL, command=self.tree.yview)
        scroll_y.grid(row=0, column=1, sticky=tk.NS)
        self.tree.configure(yscrollcommand=scroll_y.set)

        scroll_x = ttk.Scrollbar(master=self, orient=tk.HORIZONTAL, command=self.tree.xview)
        scroll_x.grid(row=1, column=0, sticky=tk.EW)
        self.tree.configure(xscrollcommand=scroll_x.set)

        self.tree.bind('<<TreeviewSelect>>', self.selecting_item)

        self.project = self.nametowidget(".").project

    def populate(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        self.project = self.nametowidget(".").load_project()

        for file_id, i in self.project["files"].items():
            self.tree.insert("", tk.END, values=(file_id, i.get("source"), i.get("path"), i.get("labels"),
                                                 dt.datetime.fromtimestamp(i.get("c_time")).strftime(
                                                     "%Y-%m-%d %H:%M:%S")))

    def selecting_item(self, event):
        for child in self.driver.detail_pan.winfo_children():
            child.destroy()

        for child in self.driver.right_pan.winfo_children():
            child.destroy()

        item_index = self.tree.focus()
        item_obj = self.tree.item(item_index)
        item_vals = item_obj.get("values")

        try:
            file_id = item_vals[0]
            path_to_screen = item_vals[2]
        except IndexError:
            return

        try:
            self.driver.detail_view = DetailPan(master=self.driver.detail_pan, driver=self.driver,
                                                item_index=item_index, file_id=file_id)

            self.driver.zoom_view = ZoomPan(master=self.driver.right_pan, path=path_to_screen)

            gc.disable()
            gc.collect()
            gc.enable()

        except tk.TclError:
            return
        except KeyError:
            return
